from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import ReportResult


SEASON_ORDER = {"Spring": 1, "Summer": 2, "Fall": 3}


def _format_course_entry(course_title: str, credits: float) -> str:
    return f"{course_title} - {float(credits):g}cr"


def _format_total_value(value: float) -> int | str:
    if not value:
        return ""
    return int(value) if float(value).is_integer() else value


def _build_preview_dataframe(grouped: pd.DataFrame, ordered_semesters: list[str]) -> pd.DataFrame:
    preview_rows: list[dict[str, object]] = []
    columns = ["Faculty Name", *ordered_semesters]

    for faculty_name, faculty_rows in grouped.groupby("Faculty Name", sort=True):
        preview_rows.append({column: column for column in columns})

        semester_entries: dict[str, list[str]] = {}
        semester_totals: dict[str, float] = {}
        max_rows = 0

        for semester in ordered_semesters:
            semester_rows = (
                faculty_rows.loc[faculty_rows["Term Label"].eq(semester), ["Course Title", "Cr."]]
                .sort_values(["Cr.", "Course Title"], ascending=[True, True])
            )
            entries = [
                _format_course_entry(course_title, credits)
                for course_title, credits in semester_rows.itertuples(index=False, name=None)
            ]
            semester_entries[semester] = entries
            semester_totals[semester] = float(semester_rows["Cr."].sum())
            max_rows = max(max_rows, len(entries))

        for row_index in range(max_rows):
            row = {"Faculty Name": faculty_name if row_index == 0 else ""}
            for semester in ordered_semesters:
                entries = semester_entries[semester]
                row[semester] = entries[row_index] if row_index < len(entries) else ""
            preview_rows.append(row)

        total_credits = sum(semester_totals.values())
        total_row = {"Faculty Name": f"total credits: {total_credits:g}"}
        for semester in ordered_semesters:
            total_row[semester] = _format_total_value(semester_totals[semester])
        preview_rows.append(total_row)
        preview_rows.append({column: "" for column in columns})
        preview_rows.append({column: "" for column in columns})

    return pd.DataFrame(preview_rows, columns=columns)


def _dataframe_to_styled_excel_bytes(dataframe: pd.DataFrame, sheet_name: str) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        worksheet = writer.book[sheet_name]
        thin = Side(style="thin")
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        total_font = Font(color="0000FF")
        total_bold_font = Font(color="0000FF", bold=True)

        for row_idx in range(1, worksheet.max_row + 1):
            row_values = [worksheet.cell(row_idx, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
            non_empty = [value for value in row_values if value not in (None, "")]
            is_header = non_empty == list(dataframe.columns)
            is_total = isinstance(worksheet.cell(row_idx, 1).value, str) and str(worksheet.cell(row_idx, 1).value).startswith(
                "total credits:"
            )

            if is_header:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(bottom=thin)
            elif is_total:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.border = Border(bottom=thin)
                    if col_idx == 1:
                        cell.font = total_bold_font
                    elif cell.value not in (None, ""):
                        cell.font = total_font
                worksheet.cell(row_idx, 1).alignment = Alignment(horizontal="right")

        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value not in (None, "") else 0
                for cell in worksheet[column_letter]
            )
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 13), 40)

    buffer.seek(0)
    return buffer.read()


def generate_credits_per_teacher_report(df: pd.DataFrame, department: str) -> ReportResult:
    working = df.copy()
    working["SEM Department"] = working["SEM Department"].astype("string").str.strip()
    working["Faculty Name"] = working["Faculty Name"].astype("string").str.strip()
    working["Subject"] = working.get("Subject", pd.Series(index=working.index, dtype="string")).astype("string").str.strip()
    working["Course Number"] = (
        working.get("Course Number", pd.Series(index=working.index, dtype="string")).astype("string").str.strip()
    )
    working["Cr."] = pd.to_numeric(working["Cr."], errors="coerce").fillna(0)
    working["Term Label"] = working["Term Label"].astype("string").str.strip()
    working["Term Year"] = pd.to_numeric(working["Term Year"], errors="coerce")
    working["Term Season"] = working["Term Season"].astype("string").str.strip()
    working["Course Title"] = (working["Subject"].fillna("") + working["Course Number"].fillna("")).str.strip()
    working["Course Title"] = working["Course Title"].replace({"": pd.NA})

    filtered = working.loc[working["SEM Department"].eq(department)].copy()
    filtered = filtered.loc[
        filtered["Faculty Name"].notna() & filtered["Term Label"].notna() & filtered["Course Title"].notna()
    ].copy()

    grouped = (
        filtered.groupby(
            ["Faculty Name", "Course Title", "Term Label", "Term Year", "Term Season"],
            dropna=False,
        )["Cr."]
        .sum()
        .reset_index()
    )
    grouped = grouped.loc[grouped["Cr."].gt(0)].copy()

    if grouped.empty:
        preview_df = pd.DataFrame(columns=["Faculty Name"])
    else:
        semester_order = (
            grouped[["Term Label", "Term Year", "Term Season"]]
            .drop_duplicates()
            .assign(season_rank=lambda x: x["Term Season"].map(SEASON_ORDER).fillna(99))
            .sort_values(["Term Year", "season_rank", "Term Label"])
        )
        ordered_semesters = semester_order["Term Label"].tolist()
        preview_df = _build_preview_dataframe(grouped, ordered_semesters)

    worksheet_name = "Credits per Teacher"
    output_filename = f"{department}_credits_per_teacher.xlsx"
    excel_bytes = _dataframe_to_styled_excel_bytes(preview_df, worksheet_name)

    return ReportResult(
        report_name=f"{department} Credits per Teacher",
        preview_df=preview_df,
        excel_bytes=excel_bytes,
        output_filename=output_filename,
        worksheet_name=worksheet_name,
        metadata={"department": department},
    )
