from __future__ import annotations

import re
from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..excel_utils import workbook_to_excel_bytes
from ..models import ReportBundle, ReportResult


DEFAULT_PROGRAM_ORDER = [
    "DIS",
    "DMIN-INT",
    "DMIN-NAD",
    "EGW-Cert",
    "MAPM-ENGL",
    "MAPM-HISP",
    "MAR-INT",
    "MAR-NAD",
    "MAYYM",
    "MDIV",
    "MDIV-FL",
    "PHD/THD",
    "SOLC",
    "Urban",
]

DEFAULT_DEPARTMENT_ORDER = [
    "CHIS",
    "DMIN",
    "DSLE",
    "GSEM",
    "MA_Religion",
    "MSSN",
    "NTST",
    "OTST",
    "PATH",
    "THST",
]


def _copy_frame(df: pd.DataFrame) -> pd.DataFrame:
    return df.copy() if df is not None else pd.DataFrame()


def _string_column(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(index=df.index, dtype="string")
    return df[column].astype("string").str.strip()


def _numeric_column(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(index=df.index, dtype="float64")
    return pd.to_numeric(df[column], errors="coerce")


def _safe_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _program_equals(series: pd.Series, program: str) -> pd.Series:
    return series.astype("string").fillna("").eq(str(program))


def _prep_working_df(df: pd.DataFrame) -> pd.DataFrame:
    working = _copy_frame(df)
    if working.empty:
        return working
    text_columns = [
        "Department",
        "SEM Department",
        "Program",
        "Faculty Name",
        "Course",
        "Catalog Title",
        "Semester",
        "Mode",
        "Location",
        "account to be charged",
        "Source Workbook",
        "Subject",
        "Course Number",
        "Course Section",
        "Term Season",
        "Term Label",
    ]
    for column in text_columns:
        working[column] = _string_column(working, column)
    working["load/contract"] = _string_column(working, "load/contract").str.lower().fillna("")
    working["Cr."] = _numeric_column(working, "Cr.").fillna(0)
    working["rate per credit"] = _numeric_column(working, "rate per credit").fillna(0)
    working["total contract amount"] = _numeric_column(working, "total contract amount").fillna(0)
    working["Fiscal Year"] = pd.to_numeric(working.get("Fiscal Year"), errors="coerce")
    working["Term Year"] = pd.to_numeric(working.get("Term Year"), errors="coerce")
    working["Department"] = working["Department"].fillna(working["SEM Department"])
    working["ORG code"] = working["account to be charged"].map(_org_code_from_account)
    working["Term Code Derived"] = working["Semester"].map(_term_code_from_semester)
    working["Course Number Text"] = _course_number_text(working)
    working["Course Text"] = _course_text(working)
    working["Credits from Load"] = working["Cr."].where(working["load/contract"].eq("load"), 0)
    working["Credits from Contract"] = working["Cr."].where(working["load/contract"].eq("contract"), 0)
    return working


def _filter_department(working: pd.DataFrame, department: str) -> pd.DataFrame:
    if working.empty:
        return working
    department_value = str(department).strip()
    mask = working["Department"].eq(department_value) | working["SEM Department"].eq(department_value)
    return working.loc[mask].copy()


def _program_columns(working: pd.DataFrame) -> list[str]:
    seen = _string_column(working, "Program").dropna().unique().tolist() if not working.empty else []
    ordered = [program for program in DEFAULT_PROGRAM_ORDER if program in seen]
    extras = sorted(program for program in seen if program not in ordered)
    return ordered + extras


def _department_order(working: pd.DataFrame) -> list[str]:
    seen = _string_column(working, "Department").dropna().unique().tolist() if not working.empty else []
    ordered = [department for department in DEFAULT_DEPARTMENT_ORDER if department in seen]
    extras = sorted(department for department in seen if department not in ordered)
    return ordered + extras


def _split_name(name: object) -> tuple[str, str]:
    text = _safe_text(name)
    if "," in text:
        last, first = text.split(",", 1)
        return last.strip(), first.strip()
    parts = text.split()
    if len(parts) >= 2:
        return parts[-1], " ".join(parts[:-1])
    return text, ""


def _org_code_from_account(value: object) -> str:
    text = _safe_text(value)
    if ":" in text:
        text = text.split(":", 1)[1].strip()
    match = re.search(r"\b(\d{4})\b", text)
    return match.group(1) if match else ""


def _term_code_from_semester(semester: object) -> str:
    text = _safe_text(semester)
    if not text:
        return ""
    parts = text.split()
    if len(parts) >= 2 and parts[-1].isdigit():
        season = parts[0].lower()
        season_code = {"spring": "1", "summer": "2", "fall": "3"}.get(season, "0")
        return f"{parts[-1]}{season_code}"
    return text


def _course_number_text(working: pd.DataFrame) -> pd.Series:
    if "Course Number" not in working.columns:
        return _string_column(working, "Course")
    values = working["Course Number"]
    return values.apply(
        lambda value: str(int(value)) if isinstance(value, float) and float(value).is_integer() else _safe_text(value)
    ).astype("string")


def _course_text(working: pd.DataFrame) -> pd.Series:
    if "Course" in working.columns:
        course = _string_column(working, "Course")
        if course.fillna("").str.len().gt(0).any():
            return course
    subject = _string_column(working, "Subject").fillna("")
    number = _course_number_text(working).fillna("")
    section = _string_column(working, "Course Section").fillna("")
    return (
        subject
        + number.where(number.eq(""), number)
        + section.map(lambda value: f"-{value}" if pd.notna(value) and value != "" else "")
    ).astype("string")


def _semester_sort_key(value: object) -> tuple[int, int]:
    text = _safe_text(value)
    if not text:
        return (9999, 999)
    parts = text.split()
    if len(parts) < 2 or not parts[-1].isdigit():
        return (9999, 999)
    season_order = {"summer": 0, "fall": 1, "spring": 2}
    return (int(parts[-1]), season_order.get(parts[0].lower(), 999))


def _ordered_semesters(series: pd.Series) -> list[str]:
    values = [str(value).strip() for value in series.dropna().tolist() if str(value).strip()]
    return sorted(dict.fromkeys(values), key=_semester_sort_key)


def _academic_year_sort_key(value: object) -> tuple[int, int]:
    text = _safe_text(value)
    if not text:
        return (9999, 999)
    parts = text.split()
    if len(parts) < 2 or not parts[-1].isdigit():
        return (9999, 999)
    season = parts[0].lower()
    year = int(parts[-1])
    if season == "spring":
        return (year - 1, 2)
    if season == "summer":
        return (year, 0)
    if season == "fall":
        return (year, 1)
    return (year, 999)


def _academic_year_semesters(series: pd.Series) -> list[str]:
    values = [str(value).strip() for value in series.dropna().tolist() if str(value).strip()]
    ordered = sorted(dict.fromkeys(values), key=_academic_year_sort_key)
    if not ordered:
        return []
    first_year = _academic_year_sort_key(ordered[0])[0]
    return [value for value in ordered if _academic_year_sort_key(value)[0] == first_year]


def _department_cycle_semesters(series: pd.Series) -> list[str]:
    values = [str(value).strip() for value in series.dropna().tolist() if str(value).strip()]
    if not values:
        return []
    starts = [_academic_year_sort_key(value)[0] for value in values]
    start_year = min(starts)
    target_labels = [f"Summer {start_year}", f"Fall {start_year}", f"Spring {start_year + 1}"]
    seen = set(values)
    return [label for label in target_labels if label in seen]


def _clean_numeric_display(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = frame.copy()
    for column in columns:
        if column in out.columns:
            out[column] = out[column].apply(
                lambda x: int(x) if isinstance(x, float) and float(x).is_integer() else x
            )
    return out


def _blank_zeroes(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = frame.copy()
    for column in columns:
        if column in out.columns:
            out[column] = out[column].apply(lambda value: pd.NA if pd.notna(value) and value == 0 else value)
    return out


def _blank_row(columns: list[str]) -> dict[str, object]:
    return {column: pd.NA for column in columns}


def _clone_sheet_into_workbook(source_ws, target_wb, title: str) -> None:
    target_ws = target_wb.create_sheet(title=title[:31] or "Sheet")
    for row in source_ws.iter_rows():
        for cell in row:
            target = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target.font = copy(cell.font)
                target.fill = copy(cell.fill)
                target.border = copy(cell.border)
                target.alignment = copy(cell.alignment)
                target.number_format = cell.number_format
                target.protection = copy(cell.protection)
            if cell.hyperlink:
                target._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                target.comment = copy(cell.comment)
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
        target_ws.column_dimensions[key].hidden = dim.hidden
    for key, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = dim.height
        target_ws.row_dimensions[key].hidden = dim.hidden
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines


def _build_contracts_overview_template(start_year: int | None) -> Workbook:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Contracts"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 3.5,
        "B": 66.83,
        "C": 25.16,
        "D": 15.0,
        "E": 13.5,
        "F": 15.5,
        "G": 11.66,
        "H": 12.16,
        "I": 9.33,
        "J": 8.83,
        "K": 13.0,
        "L": 13.0,
        "M": 13.0,
    }
    for key, value in widths.items():
        ws.column_dimensions[key].width = value

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    header_font = Font(name="Arial", size=11, bold=True)
    body_font = Font(name="Arial", size=11)
    title_font = Font(name="Arial", size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    currency_fill = PatternFill(fill_type="solid", fgColor="CCFFFF")
    light_blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    olive_fill = PatternFill(fill_type="solid", fgColor="D8E4BC")

    def merge(range_ref: str) -> None:
        ws.merge_cells(range_ref)

    def set_text(cell_ref: str, value, *, font=body_font, alignment=None):
        cell = ws[cell_ref]
        cell.value = value
        cell.font = copy(font)
        if alignment is not None:
            cell.alignment = alignment

    def fill_range(row: int, col_start: int, col_end: int, fill) -> None:
        for col in range(col_start, col_end + 1):
            ws.cell(row, col).fill = copy(fill)

    def fill_table_columns(row_start: int, row_end: int) -> None:
        white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        for row in range(row_start, row_end + 1):
            for col in range(2, 6):
                ws.cell(row, col).fill = copy(white_fill)
            ws.cell(row, 6).fill = copy(light_blue_fill)
            ws.cell(row, 7).fill = copy(olive_fill)
            ws.cell(row, 8).fill = copy(olive_fill)

    def box(row: int, col_start: int, col_end: int, *, left_side=thin, right_side=thin, top_side=thin, bottom_side=thin):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            cell.border = Border(
                left=left_side if col == col_start else thin,
                right=right_side if col == col_end else thin,
                top=top_side,
                bottom=bottom_side,
            )

    budget_end_year = start_year + 1 if start_year is not None else None
    budget_label = f"May {start_year}-April {budget_end_year}" if start_year is not None else "May - April"
    budget_label_spaced = f"May {start_year} - April {budget_end_year}" if start_year is not None else "May - April"
    summer_label = f"Summer {start_year}" if start_year is not None else "Summer"
    fall_label = f"Fall {start_year}" if start_year is not None else "Fall"
    spring_label = f"Spring {budget_end_year}" if budget_end_year is not None else "Spring"

    for rng in [
        "B2:F2", "B3:F3", "B4:F4", "B5:F5",
        "B7:E7", "B8:E8", "B9:E9", "B10:E10",
        "B12:B13", "C12:C13", "D12:E13", "F12:F13",
        "D14:E14", "D15:E15", "D16:E16", "D17:E17", "D18:E18",
        "B19:E19", "B20:E20", "B21:E21", "B22:E22",
        "B24:H24", "B25:H25", "B26:H26", "B27:H27",
        "B55:C55", "B93:C93", "B94:C94",
    ]:
        merge(rng)

    for rng in [
        "J57:L57", "J58:L58", "J59:L59", "J60:L60", "J61:L61", "J62:L62",
        "J63:L63", "J64:L64", "J65:L65", "J66:L66", "J67:L67", "J68:L68",
        "J69:L69", "J70:L70", "J71:L71",
    ]:
        merge(rng)

    set_text("B2", "ANDREWS UNIVERSITY", font=title_font, alignment=center)
    set_text("B3", "Budget Request", font=title_font, alignment=center)
    set_text("B4", budget_label, font=title_font, alignment=center)
    set_text("B5", "SCHEDULE OF DETAILS", font=title_font, alignment=center)
    set_text("B7", f"Credits on {summer_label} schedule", font=header_font, alignment=left)
    set_text("B8", f"Credits on {fall_label} schedule", font=header_font, alignment=left)
    set_text("B9", f"Credits on {spring_label} schedule", font=header_font, alignment=left)
    set_text("B10", "Total proposed credits offered for the year", font=header_font, alignment=left)
    ws["F10"] = "=SUM(F7:F9)"

    for row in [7, 8, 9]:
        box(row, 2, 6)
        ws.cell(row, 6).font = copy(header_font)
        ws.cell(row, 6).alignment = center
        ws.cell(row, 6).number_format = "0"
    box(10, 2, 6, left_side=thin, right_side=medium, top_side=thin, bottom_side=medium)
    ws["F10"].font = copy(header_font)
    ws["F10"].alignment = center

    set_text("B12", "Salaried Faculty Name", font=header_font, alignment=center)
    set_text("C12", "Load Reduction Reason", font=header_font, alignment=center)
    set_text("D12", "Load by Policy", font=header_font, alignment=center)
    set_text("F12", "Teaching Load", font=header_font, alignment=center)
    fill_range(12, 2, 6, light_blue_fill)
    fill_range(13, 2, 6, light_blue_fill)
    for row in range(12, 19):
        box(row, 2, 8, left_side=thin, right_side=medium, top_side=thin, bottom_side=thin)
        for col in range(2, 9):
            ws.cell(row, col).font = copy(body_font)
        for col in [7, 8]:
            ws.cell(row, col).border = Border()
    for ref in ["B12", "C12", "D12", "F12"]:
        ws[ref].font = copy(header_font)
    set_text("B14", "Instructor1")
    set_text("C14", "Dept chair")
    set_text("B15", "Instructor2")
    set_text("B16", "Instructor3")
    set_text("B17", "Instructor4")
    set_text("B18", "Instructor5")
    set_text("B20", "Proposed credits to be taught by Seminary faculty", font=header_font, alignment=left)
    set_text("B21", "Proposed credits to be taught by contract", font=header_font, alignment=left)
    set_text("B22", "TOTAL", font=header_font, alignment=left)
    ws["F20"] = "=SUM(F14:F17)"
    ws["F21"] = "=+D94"
    ws["F22"] = "=SUM(F20:F21)"
    for row in [20, 21]:
        box(row, 2, 6)
    box(22, 2, 6, left_side=thin, right_side=medium, top_side=thin, bottom_side=medium)

    set_text("B24", "ANDREWS UNIVERSITY", font=title_font, alignment=center)
    set_text("B25", "OT Budget Request", font=title_font, alignment=center)
    set_text("B26", budget_label_spaced, font=title_font, alignment=center)
    set_text("B27", "SCHEDULE OF DETAILS", font=title_font, alignment=center)
    set_text("B29", "Schedule 1(b) -- Contract Salaries", font=title_font)
    set_text("B30", "Persons working less than half time for ", font=header_font)
    set_text("B31", "university and/or employed on a ", font=header_font)
    set_text("B32", "contract basis, including retirees", font=header_font)
    set_text("C30", "Academic Rank", font=header_font, alignment=center)
    set_text("C31", "or Equivalent", font=header_font, alignment=center)
    set_text("D30", "Number of", font=header_font, alignment=center)
    set_text("D31", "Credits Taught", font=header_font, alignment=center)
    set_text("E30", "Rate per", font=header_font, alignment=center)
    set_text("E31", "Credit", font=header_font, alignment=center)
    set_text("F30", "Contract", font=header_font, alignment=center)
    set_text("F31", "Amount", font=header_font, alignment=center)
    set_text("G30", "Salary", font=header_font, alignment=center)
    set_text("G31", "Contracts", font=header_font, alignment=center)
    set_text("H30", "PhD", font=header_font, alignment=center)
    set_text("H31", "Contracts", font=header_font, alignment=center)
    ws["G32"] = 9120
    ws["H32"] = 9260
    fill_table_columns(30, 94)
    for row in [30, 31, 32]:
        box(row, 2, 8, left_side=medium if row == 32 else thin, right_side=medium, top_side=thin, bottom_side=thin)

    def contract_section_header(row: int, title: str) -> None:
        set_text(f"B{row}", title, font=header_font)
        box(row, 2, 8, left_side=medium, right_side=medium, top_side=medium if row in {33, 57, 67, 78, 85} else thin, bottom_side=medium if row in {57, 67, 78, 85} else thin)

    def semester_row(row: int, label: str) -> None:
        set_text(f"B{row}", label, font=header_font)
        box(row, 2, 8, left_side=medium, right_side=medium, top_side=thin, bottom_side=thin)

    def detail_row(row: int) -> None:
        box(row, 2, 8, left_side=thin, right_side=medium, top_side=thin, bottom_side=thin)
        ws[f"F{row}"] = f"=E{row}*D{row}"
        ws[f"D{row}"].number_format = '0.##'
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"F{row}"].number_format = '"$"#,##0.00'
        ws[f"G{row}"].number_format = '"$"#,##0.00'
        ws[f"H{row}"].number_format = '"$"#,##0.00'

    contract_section_header(33, "ON-CAMPUS CONTRACTS")
    semester_row(34, summer_label)
    for row in [35, 36, 37]:
        detail_row(row)
    semester_row(39, fall_label)
    for row in [40, 41, 42, 43, 44, 45]:
        detail_row(row)
    semester_row(47, spring_label)
    for row in [48, 49, 50, 51, 52, 53]:
        detail_row(row)
    set_text("B55", "Total On Campus Contracts", font=header_font, alignment=right)
    ws["D55"] = "=SUM(D35:D54)"
    ws["F55"] = "=SUM(F35:F54)"
    ws["G55"] = "=SUM(G35:G54)"
    ws["H55"] = "=SUM(H35:H54)"
    box(55, 2, 8, left_side=medium, right_side=medium, top_side=medium, bottom_side=medium)

    contract_section_header(57, "MAPM - ENG CONTRACT")
    semester_row(59, fall_label)
    for row in [60]:
        detail_row(row)
    semester_row(62, spring_label)
    for row in [63]:
        detail_row(row)

    contract_section_header(67, "MAPM - HISP CONTRACT")
    semester_row(68, summer_label)
    for row in [69]:
        detail_row(row)
    semester_row(71, fall_label)
    for row in [72, 73]:
        detail_row(row)
    semester_row(75, spring_label)
    for row in [76]:
        detail_row(row)

    contract_section_header(78, "MA RELIGION CONTRACT")
    semester_row(80, fall_label)
    for row in [81, 82]:
        detail_row(row)

    contract_section_header(85, "MDIV Spanish FL Cohort")
    semester_row(87, fall_label)
    for row in [88]:
        detail_row(row)

    set_text("B93", "Total Extension School Classes", font=header_font, alignment=right)
    ws["D93"] = "=SUM(D58:D92)"
    ws["F93"] = "=SUM(F58:F92)"
    ws["G93"] = "=SUM(G58:G92)"
    ws["H93"] = "=SUM(H58:H92)"
    box(93, 2, 8, left_side=medium, right_side=medium, top_side=medium, bottom_side=medium)

    set_text("B94", "Total Contracts", font=header_font, alignment=right)
    ws["D94"] = "=+D55+D93"
    ws["F94"] = "=+F55+F93"
    ws["G94"] = "=+G55+G93"
    ws["H94"] = "=+H55+H93"
    ws["F94"].number_format = '"$"#,##0.00'
    ws["G94"].number_format = '"$"#,##0.00'
    ws["H94"].number_format = '"$"#,##0.00'
    box(94, 2, 8, left_side=medium, right_side=medium, top_side=medium, bottom_side=medium)

    for row in range(33, 95):
        if row not in {33, 55, 57, 67, 78, 85, 93, 94}:
            box(row, 2, 8, left_side=thin, right_side=medium, top_side=thin, bottom_side=thin)

    location_labels = [
        "LOCATION ABBREVIATIONS",
        "Columbia Union - CU",
        "Lake Union - LU",
        "Mid-America Union - MAU",
        "Pacific Union - PU",
        "North Pacific Union - NPU",
        "Poland/Baltic States - PolBal",
        "Poland/Baltic States - PolBal",
        "Poland/Baltic States - PolBal",
        "Romania - Rom",
        "Ukraine - Ukr",
        "Main Campus - MC",
        "Atlantic Union - AtlU",
        "Hong Kong - HK",
        "Lake Union - LU",
    ]
    location_rows = [57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 68, 69, 74, 83, 71]
    for row, label in zip(location_rows, location_labels):
        merge(f"J{row}:L{row}")
        set_text(f"J{row}", label, font=body_font if row != 57 else header_font, alignment=center if row != 57 else left)

    return workbook


def _safe_sum(series: pd.Series) -> float:
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())


def _department_org_code(df: pd.DataFrame) -> str:
    if df.empty or "ORG code" not in df.columns:
        return ""
    non_blank = df["ORG code"].dropna().astype("string").str.strip()
    non_blank = non_blank[non_blank.ne("")]
    return str(non_blank.iloc[0]) if not non_blank.empty else ""


def _faculty_program_matrix(
    df: pd.DataFrame,
    value_column: str,
    programs: list[str],
    include_total: bool = True,
) -> pd.DataFrame:
    if df.empty:
        columns = ["Faculty Name", *programs]
        if include_total:
            columns.append("Total")
        return pd.DataFrame(columns=columns)
    pivot = (
        pd.pivot_table(
            df,
            index="Faculty Name",
            columns="Program",
            values=value_column,
            aggfunc="sum",
            fill_value=0,
        )
        .reindex(columns=programs, fill_value=0)
        .reset_index()
    )
    if include_total:
        pivot["Total"] = pivot[programs].sum(axis=1)
    return pivot


def _report_001a_columns(program_cols: list[str]) -> list[str]:
    return [
        "Department",
        "ORG code",
        "Mode",
        "Semester",
        "LastName",
        "FirstName",
        "CourseNumber",
        "Cr.",
        "Cost",
        *program_cols,
        "Check Total",
        "Notes",
    ]


def _overview_report_001a(working: pd.DataFrame) -> pd.DataFrame:
    program_cols = _program_columns(working)
    columns = _report_001a_columns(program_cols)
    if working.empty:
        return pd.DataFrame(columns=columns)

    report = working.copy()
    names = report["Faculty Name"].map(_split_name)
    report["LastName"] = names.map(lambda value: value[0])
    report["FirstName"] = names.map(lambda value: value[1])
    report["CourseNumber"] = report["Course Text"]
    report["Cost"] = report["total contract amount"]
    for program in program_cols:
        report[program] = report["Cost"].where(_program_equals(report["Program"], program), 0)
    report["Check Total"] = report[program_cols].sum(axis=1) if program_cols else report["Cost"]
    report["Notes"] = pd.NA
    report = report.sort_values(
        ["Department", "Semester", "LastName", "FirstName", "CourseNumber"],
        key=lambda col: col.map(_semester_sort_key) if col.name == "Semester" else col,
        na_position="last",
    )

    lines: list[dict[str, object]] = []
    for department in _department_order(report):
        dept_df = report.loc[report["Department"].eq(department), columns].copy()
        if dept_df.empty:
            continue
        lines.extend(dept_df.to_dict("records"))
        summary = {column: pd.NA for column in columns}
        summary["Department"] = department
        summary["ORG code"] = _department_org_code(dept_df)
        summary["LastName"] = "DEPARTMENT TOTAL"
        for column in ["Cr.", "Cost", "Check Total", *program_cols]:
            summary[column] = _safe_sum(dept_df[column])
        lines.append(summary)
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["Cr.", "Cost", "Check Total", *program_cols])


def _overview_report_001b(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Faculty", "Course", "Mode", "Location", "Credits", "Rate", "Total"]
    if working.empty:
        return pd.DataFrame(columns=columns)

    contract_rows = working.loc[working["load/contract"].eq("contract")].copy()
    if contract_rows.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(contract_rows):
        dept_df = contract_rows.loc[contract_rows["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        lines.append({"Faculty": department})
        for program in _program_columns(dept_df):
            program_df = dept_df.loc[dept_df["Program"].eq(program)].copy()
            if program_df.empty:
                continue
            lines.append({"Faculty": f"{program} contracts"})
            for semester in _ordered_semesters(program_df["Semester"]):
                semester_df = program_df.loc[program_df["Semester"].eq(semester)].copy()
                lines.append({"Faculty": semester})
                semester_df = semester_df.sort_values(["Faculty Name", "Course Text"], na_position="last")
                for _, row in semester_df.iterrows():
                    lines.append(
                        {
                            "Faculty": row["Faculty Name"],
                            "Course": row["Course Text"],
                            "Mode": row["Mode"],
                            "Location": row["Location"],
                            "Credits": row["Cr."],
                            "Rate": row["rate per credit"],
                            "Total": row["total contract amount"],
                        }
                    )
                lines.append(_blank_row(columns))
            lines.append(
                {
                    "Faculty": f"Total {program}",
                    "Credits": _safe_sum(program_df["Cr."]),
                    "Total": _safe_sum(program_df["total contract amount"]),
                }
            )
            lines.append(_blank_row(columns))
        lines.append(
            {
                "Faculty": "Total contracts",
                "Credits": _safe_sum(dept_df["Cr."]),
                "Total": _safe_sum(dept_df["total contract amount"]),
            }
        )
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["Credits", "Rate", "Total"])


def _overview_report_001c(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Faculty", "Course", "Program", "LoadType", "Mode", "Location", "Credits", "Rate", "Total"]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        lines.append({"Faculty": department})
        for semester in _ordered_semesters(dept_df["Semester"]):
            semester_df = dept_df.loc[dept_df["Semester"].eq(semester)].copy()
            if semester_df.empty:
                continue
            lines.append({"Faculty": semester})
            semester_df = semester_df.sort_values(["Faculty Name", "Course Text"], na_position="last")
            for _, row in semester_df.iterrows():
                lines.append(
                    {
                        "Faculty": row["Faculty Name"],
                        "Course": row["Course Text"],
                        "Program": row["Program"],
                        "LoadType": _safe_text(row["load/contract"]).title(),
                        "Mode": row["Mode"],
                        "Location": row["Location"],
                        "Credits": row["Cr."],
                        "Rate": row["rate per credit"] if _is_contract_value(row["load/contract"]) else 0,
                        "Total": row["total contract amount"] if _is_contract_value(row["load/contract"]) else 0,
                    }
                )
            lines.append(
                {
                    "Faculty": f"Total {semester}",
                    "Credits": _safe_sum(semester_df["Cr."]),
                    "Total": _safe_sum(semester_df["total contract amount"]),
                }
            )
            lines.append(_blank_row(columns))
        lines.append(
            {
                "Faculty": "Total contracts",
                "Credits": _safe_sum(dept_df["Cr."]),
                "Total": _safe_sum(dept_df["total contract amount"]),
            }
        )
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["Credits", "Rate", "Total"])


def _overview_report_001f(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Department", "Faculty", "Course", "Program", "LoadType", "Mode", "Location", "Credits", "Rate", "Total"]
    if working.empty:
        return pd.DataFrame(columns=columns)

    department = str(_string_column(working, "Department").dropna().iloc[0]) if not working.empty else ""
    lines: list[dict[str, object]] = []
    faculty_names = sorted(_string_column(working, "Faculty Name").dropna().unique().tolist())

    first_faculty = True
    for faculty in faculty_names:
        faculty_df = working.loc[working["Faculty Name"].eq(faculty)].copy()
        if faculty_df.empty:
            continue
        faculty_total_credits = 0.0
        faculty_total_amount = 0.0
        first_semester_row = True
        for semester in _department_cycle_semesters(faculty_df["Semester"]):
            semester_df = faculty_df.loc[faculty_df["Semester"].eq(semester)].copy()
            if semester_df.empty:
                continue
            lines.append(
                {
                    "Department": department if first_faculty and first_semester_row else pd.NA,
                    "Faculty": faculty if first_semester_row else pd.NA,
                    "Course": semester,
                }
            )
            semester_total_credits = 0.0
            semester_total_amount = 0.0
            semester_df = semester_df.sort_values(["Course Text", "Mode", "Location"], na_position="last")
            for _, row in semester_df.iterrows():
                semester_total_credits += float(row["Cr."])
                semester_total_amount += float(row["total contract amount"])
                lines.append(
                    {
                        "Course": row["Course Text"],
                        "Program": row["Program"],
                        "LoadType": _safe_text(row["load/contract"]).title(),
                        "Mode": row["Mode"],
                        "Location": row["Location"],
                        "Credits": row["Cr."],
                        "Rate": row["rate per credit"] if _is_contract_value(row["load/contract"]) else 0,
                        "Total": row["total contract amount"] if _is_contract_value(row["load/contract"]) else 0,
                    }
                )
            lines.append({"Course": f"Total {semester}", "Credits": semester_total_credits, "Total": semester_total_amount})
            faculty_total_credits += semester_total_credits
            faculty_total_amount += semester_total_amount
            first_semester_row = False
        lines.append({"Course": "TOTAL", "Credits": faculty_total_credits, "Total": faculty_total_amount})
        first_faculty = False
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["Credits", "Rate", "Total"])


def _overview_report_001f_all_departments(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Department", "Faculty", "Course", "Program", "LoadType", "Mode", "Location", "Credits", "Rate", "Total"]
    if working.empty:
        return pd.DataFrame(columns=columns)
    frames = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        frames.append(_overview_report_001f(dept_df))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=columns)


def _report_002a_columns(program_cols: list[str]) -> list[str]:
    base_columns = [
        "Department",
        "ORG code",
        "Account Charged",
        "Load/Contract",
        "Term Code",
        "Mode/Cohort",
        "Semester",
        "Faculty Name",
        "Subject",
        "Course Number",
        "Course Section",
        "Credits from Load",
        "Credits from Contract",
        "Cr.",
        "Cost",
    ]
    pairs = []
    for program in program_cols:
        pairs.extend([f"{program} Credits", f"{program} Cost"])
    return base_columns + pairs


def _overview_report_002a(working: pd.DataFrame) -> pd.DataFrame:
    program_cols = _program_columns(working)
    columns = _report_002a_columns(program_cols)
    if working.empty:
        return pd.DataFrame(columns=columns)

    report = working.copy()
    report["Account Charged"] = report["account to be charged"]
    report["Load/Contract"] = report["load/contract"]
    report["Term Code"] = report["Term Code"] if "Term Code" in report.columns else report["Term Code Derived"]
    report["Mode/Cohort"] = report["Mode"]
    report["Cost"] = report["total contract amount"].where(report["load/contract"].eq("contract"), 0)
    for program in program_cols:
        program_mask = _program_equals(report["Program"], program)
        report[f"{program} Credits"] = report["Cr."].where(program_mask, 0)
        report[f"{program} Cost"] = report["Cost"].where(program_mask, 0)
    report = report[columns].sort_values(
        ["Department", "Semester", "Faculty Name", "Subject", "Course Number"],
        key=lambda col: col.map(_semester_sort_key) if col.name == "Semester" else col,
        na_position="last",
    )

    lines: list[dict[str, object]] = []
    numeric_columns = ["Credits from Load", "Credits from Contract", "Cr.", "Cost", *[f"{program} Credits" for program in program_cols], *[f"{program} Cost" for program in program_cols]]
    for department in _department_order(report):
        dept_df = report.loc[report["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        lines.extend(dept_df.to_dict("records"))
        summary = {column: pd.NA for column in columns}
        summary["Department"] = department
        summary["ORG code"] = _department_org_code(dept_df)
        summary["Faculty Name"] = "DEPARTMENT TOTAL"
        for column in numeric_columns:
            summary[column] = _safe_sum(dept_df[column])
        lines.append(summary)
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), numeric_columns)


def _overview_report_003a(working: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Department",
        "Contract Cost",
        "Load_credits",
        "Contract_credits",
        "Load/Contract Ratio",
        "Sum of all Credits",
        "Sum of all Credits wo Lilly",
        "Sum of all Credits wo NADEI's MDIV courses",
        "Load_credits wo NADEI's MDIV courses",
        "Load/Contract Ratio wo NADEI's MDIV courses",
        "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)",
        "Contract Credits wo Lilly funded Courses",
    ]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        load_credits = _safe_sum(dept_df.loc[dept_df["load/contract"].eq("load"), "Cr."])
        contract_credits = _safe_sum(dept_df.loc[dept_df["load/contract"].eq("contract"), "Cr."])
        contract_cost = _safe_sum(dept_df.loc[dept_df["load/contract"].eq("contract"), "total contract amount"])
        total_credits = load_credits + contract_credits
        # The current DonSheet does not expose Lilly/NADEI flags in a stable way, so these
        # legacy columns currently mirror the full totals until those markers are formalized.
        lines.append(
            {
                "Department": department,
                "Contract Cost": contract_cost,
                "Load_credits": load_credits,
                "Contract_credits": contract_credits,
                "Load/Contract Ratio": (contract_credits / load_credits) if load_credits else pd.NA,
                "Sum of all Credits": total_credits,
                "Sum of all Credits wo Lilly": total_credits,
                "Sum of all Credits wo NADEI's MDIV courses": total_credits,
                "Load_credits wo NADEI's MDIV courses": load_credits,
                "Load/Contract Ratio wo NADEI's MDIV courses": (contract_credits / load_credits) if load_credits else pd.NA,
                "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)": total_credits,
                "Contract Credits wo Lilly funded Courses": contract_credits,
            }
        )
    report = pd.DataFrame(lines, columns=columns)
    totals = {
        "Department": "TOTAL",
        "Contract Cost": _safe_sum(report["Contract Cost"]),
        "Load_credits": _safe_sum(report["Load_credits"]),
        "Contract_credits": _safe_sum(report["Contract_credits"]),
        "Sum of all Credits": _safe_sum(report["Sum of all Credits"]),
        "Sum of all Credits wo Lilly": _safe_sum(report["Sum of all Credits wo Lilly"]),
        "Sum of all Credits wo NADEI's MDIV courses": _safe_sum(report["Sum of all Credits wo NADEI's MDIV courses"]),
        "Load_credits wo NADEI's MDIV courses": _safe_sum(report["Load_credits wo NADEI's MDIV courses"]),
        "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)": _safe_sum(report["Sum of all Credits wo (Lilly + NADEI's Mdiv courses)"]),
        "Contract Credits wo Lilly funded Courses": _safe_sum(report["Contract Credits wo Lilly funded Courses"]),
    }
    totals["Load/Contract Ratio"] = totals["Contract_credits"] / totals["Load_credits"] if totals["Load_credits"] else pd.NA
    totals["Load/Contract Ratio wo NADEI's MDIV courses"] = (
        totals["Contract Credits wo Lilly funded Courses"] / totals["Load_credits wo NADEI's MDIV courses"]
        if totals["Load_credits wo NADEI's MDIV courses"]
        else pd.NA
    )
    report = pd.concat([report, pd.DataFrame([totals])], ignore_index=True)
    return _clean_numeric_display(
        report,
        [
            "Contract Cost",
            "Load_credits",
            "Contract_credits",
            "Sum of all Credits",
            "Sum of all Credits wo Lilly",
            "Sum of all Credits wo NADEI's MDIV courses",
            "Load_credits wo NADEI's MDIV courses",
            "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)",
            "Contract Credits wo Lilly funded Courses",
        ],
    )


def _budget_report_002b(working: pd.DataFrame) -> pd.DataFrame:
    program_cols = _program_columns(working)
    credit_columns: list[str] = []
    for program in program_cols:
        credit_columns.append(f"{program} Credits")
    columns = [
        "Department",
        "ORG code",
        "Course",
        "Faculty Name",
        "Semester",
        "load/contract",
        "Program",
        "Cr.",
        *credit_columns,
        "Total Credits",
    ]
    if working.empty:
        return pd.DataFrame(columns=columns)

    report = working.copy()
    for program in program_cols:
        report[f"{program} Credits"] = report["Cr."].where(_program_equals(report["Program"], program), 0)
    report["Total Credits"] = report["Cr."]
    report = report[columns].sort_values(
        ["Department", "Program", "Semester", "Course", "Faculty Name"],
        key=lambda col: col.map(_semester_sort_key) if col.name == "Semester" else col,
        na_position="last",
    )

    lines: list[dict[str, object]] = []
    numeric_columns = ["Cr.", *credit_columns, "Total Credits"]
    for department in _department_order(report):
        dept_df = report.loc[report["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        lines.extend(dept_df.to_dict("records"))
        summary = {column: pd.NA for column in columns}
        summary["Department"] = department
        summary["ORG code"] = _department_org_code(dept_df)
        summary["Course"] = "DEPARTMENT TOTAL"
        for column in numeric_columns:
            summary[column] = _safe_sum(dept_df[column])
        lines.append(summary)
        lines.append(_blank_row(columns))
    out = _clean_numeric_display(pd.DataFrame(lines, columns=columns), numeric_columns)
    return _blank_zeroes(out, ["Cr.", *credit_columns, "Total Credits"])


def _budget_report_003b(working: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Subject",
        "Contract Cost",
        "Load_credits",
        "Contract_credits",
        "Load/Contract Ratio",
        "Sum of all Credits",
        "Sum of all Credits wo Lilly",
        "Sum of all Credits wo NADEI's MDIV courses",
        "Load_credits wo NADEI's MDIV courses",
        "Load/Contract Ratio wo NADEI's MDIV courses",
        "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)",
        "Contract Credits wo Lilly funded Courses",
    ]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    subjects = sorted(_string_column(working, "Subject").dropna().unique().tolist())
    for subject in subjects:
        subject_df = working.loc[working["Subject"].eq(subject)].copy()
        load_credits = _safe_sum(subject_df.loc[subject_df["load/contract"].eq("load"), "Cr."])
        contract_credits = _safe_sum(subject_df.loc[subject_df["load/contract"].eq("contract"), "Cr."])
        contract_cost = _safe_sum(subject_df.loc[subject_df["load/contract"].eq("contract"), "total contract amount"])
        total_credits = load_credits + contract_credits
        lines.append(
            {
                "Subject": subject,
                "Contract Cost": contract_cost,
                "Load_credits": load_credits,
                "Contract_credits": contract_credits,
                "Load/Contract Ratio": (contract_credits / load_credits) if load_credits else pd.NA,
                "Sum of all Credits": total_credits,
                "Sum of all Credits wo Lilly": total_credits,
                "Sum of all Credits wo NADEI's MDIV courses": total_credits,
                "Load_credits wo NADEI's MDIV courses": load_credits,
                "Load/Contract Ratio wo NADEI's MDIV courses": (contract_credits / load_credits) if load_credits else pd.NA,
                "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)": total_credits,
                "Contract Credits wo Lilly funded Courses": contract_credits,
            }
        )
    report = pd.DataFrame(lines, columns=columns)
    totals = {
        "Subject": "TOTAL",
        "Contract Cost": _safe_sum(report["Contract Cost"]),
        "Load_credits": _safe_sum(report["Load_credits"]),
        "Contract_credits": _safe_sum(report["Contract_credits"]),
        "Sum of all Credits": _safe_sum(report["Sum of all Credits"]),
        "Sum of all Credits wo Lilly": _safe_sum(report["Sum of all Credits wo Lilly"]),
        "Sum of all Credits wo NADEI's MDIV courses": _safe_sum(report["Sum of all Credits wo NADEI's MDIV courses"]),
        "Load_credits wo NADEI's MDIV courses": _safe_sum(report["Load_credits wo NADEI's MDIV courses"]),
        "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)": _safe_sum(report["Sum of all Credits wo (Lilly + NADEI's Mdiv courses)"]),
        "Contract Credits wo Lilly funded Courses": _safe_sum(report["Contract Credits wo Lilly funded Courses"]),
    }
    totals["Load/Contract Ratio"] = (
        totals["Contract_credits"] / totals["Load_credits"] if totals["Load_credits"] else pd.NA
    )
    totals["Load/Contract Ratio wo NADEI's MDIV courses"] = (
        totals["Contract Credits wo Lilly funded Courses"] / totals["Load_credits wo NADEI's MDIV courses"]
        if totals["Load_credits wo NADEI's MDIV courses"]
        else pd.NA
    )
    report = pd.concat([report, pd.DataFrame([totals])], ignore_index=True)
    return _clean_numeric_display(
        report,
        [
            "Contract Cost",
            "Load_credits",
            "Contract_credits",
            "Sum of all Credits",
            "Sum of all Credits wo Lilly",
            "Sum of all Credits wo NADEI's MDIV courses",
            "Load_credits wo NADEI's MDIV courses",
            "Sum of all Credits wo (Lilly + NADEI's Mdiv courses)",
            "Contract Credits wo Lilly funded Courses",
        ],
    )


def _budget_report_004a(working: pd.DataFrame) -> pd.DataFrame:
    programs = _program_columns(working)
    columns = ["Department", "Faculty Name", *programs, "Total"]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        faculty_matrix = _faculty_program_matrix(dept_df, "Cr.", programs)
        for _, row in faculty_matrix.sort_values("Faculty Name", na_position="last").iterrows():
            line = {"Department": department, "Faculty Name": row["Faculty Name"]}
            for program in programs:
                line[program] = row[program]
            line["Total"] = row["Total"]
            lines.append(line)
        total_line = {"Department": department, "Faculty Name": "DEPARTMENT TOTAL"}
        for program in programs:
            total_line[program] = _safe_sum(dept_df.loc[dept_df["Program"].eq(program), "Cr."])
        total_line["Total"] = sum(total_line[program] for program in programs)
        lines.append(total_line)
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    out = _clean_numeric_display(pd.DataFrame(lines, columns=columns), [*programs, "Total"])
    return _blank_zeroes(out, [*programs, "Total"])


def _budget_report_004b(working: pd.DataFrame) -> pd.DataFrame:
    programs = _program_columns(working)
    columns = ["Department", "Faculty Name", "Total Load per person", *programs]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        load_df = dept_df.loc[dept_df["load/contract"].eq("load")].copy()
        faculty_matrix = _faculty_program_matrix(load_df, "Cr.", programs, include_total=False)
        faculty_matrix["Total Load per person"] = faculty_matrix[programs].sum(axis=1) if not faculty_matrix.empty else pd.Series(dtype="float64")
        for _, row in faculty_matrix.sort_values("Faculty Name", na_position="last").iterrows():
            line = {"Department": department, "Faculty Name": row["Faculty Name"], "Total Load per person": row["Total Load per person"]}
            for program in programs:
                line[program] = row[program]
            lines.append(line)

        def _summary_row(label: str, subset: pd.DataFrame, total_col: str) -> dict[str, object]:
            line = {"Department": department, "Faculty Name": label, "Total Load per person": _safe_sum(subset[total_col])}
            for program in programs:
                line[program] = _safe_sum(subset.loc[subset["Program"].eq(program), total_col])
            return line

        lines.append(_summary_row("DEPARTMENT TOTAL Load Credits", load_df, "Cr."))
        contract_df = dept_df.loc[dept_df["load/contract"].eq("contract")].copy()
        lines.append(_summary_row("DEPARTMENT TOTAL Contract Credits", contract_df, "Cr."))
        lines.append(_summary_row("DEPARTMENT TOTAL Contract+Load Credits", dept_df, "Cr."))
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    out = _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["Total Load per person", *programs])
    return _blank_zeroes(out, ["Total Load per person", *programs])


def _budget_report_005(working: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Program",
        "credits from Load",
        "credits from Contract",
        "credits from neither Load nor Contract",
        "Total credits",
        "Contract/Load Ratio",
    ]
    if working.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    programs = _program_columns(working)
    for program in programs:
        program_df = working.loc[working["Program"].eq(program)].copy()
        load_credits = _safe_sum(program_df.loc[program_df["load/contract"].eq("load"), "Cr."])
        contract_credits = _safe_sum(program_df.loc[program_df["load/contract"].eq("contract"), "Cr."])
        neither_credits = _safe_sum(program_df.loc[~program_df["load/contract"].isin(["load", "contract"]), "Cr."])
        total_credits = load_credits + contract_credits + neither_credits
        lines.append(
            {
                "Program": program,
                "credits from Load": load_credits,
                "credits from Contract": contract_credits,
                "credits from neither Load nor Contract": neither_credits,
                "Total credits": total_credits,
                "Contract/Load Ratio": (contract_credits / load_credits) if load_credits else pd.NA,
            }
        )
    report = pd.DataFrame(lines, columns=columns)
    totals = {
        "Program": "TOTAL",
        "credits from Load": _safe_sum(report["credits from Load"]),
        "credits from Contract": _safe_sum(report["credits from Contract"]),
        "credits from neither Load nor Contract": _safe_sum(report["credits from neither Load nor Contract"]),
        "Total credits": _safe_sum(report["Total credits"]),
    }
    totals["Contract/Load Ratio"] = (
        totals["credits from Contract"] / totals["credits from Load"] if totals["credits from Load"] else pd.NA
    )
    report = pd.concat([report, pd.DataFrame([totals])], ignore_index=True)
    return _clean_numeric_display(
        report,
        ["credits from Load", "credits from Contract", "credits from neither Load nor Contract", "Total credits"],
    )


def _budget_report_006(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Department", "Program", "contract_count", "contract_credits", "contract_dollars"]
    if working.empty:
        return pd.DataFrame(columns=columns)

    contract_df = working.loc[working["load/contract"].eq("contract")].copy()
    if contract_df.empty:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(contract_df):
        dept_df = contract_df.loc[contract_df["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        for program in _program_columns(dept_df):
            program_df = dept_df.loc[dept_df["Program"].eq(program)].copy()
            if program_df.empty:
                continue
            lines.append(
                {
                    "Department": department,
                    "Program": program,
                    "contract_count": int(len(program_df)),
                    "contract_credits": _safe_sum(program_df["Cr."]),
                    "contract_dollars": _safe_sum(program_df["total contract amount"]),
                }
            )
        lines.append(
            {
                "Department": department,
                "Program": "DEPARTMENT TOTAL",
                "contract_count": int(len(dept_df)),
                "contract_credits": _safe_sum(dept_df["Cr."]),
                "contract_dollars": _safe_sum(dept_df["total contract amount"]),
            }
        )
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), ["contract_count", "contract_credits", "contract_dollars"])


def _fiscal_years(working: pd.DataFrame) -> list[int]:
    years = [int(value) for value in working["Fiscal Year"].dropna().tolist()] if not working.empty else []
    return sorted(dict.fromkeys(years))


def _year_columns(years: list[int], prefix: str = "FY") -> list[str]:
    return [f"{prefix}{year}" for year in years]


def _budget_report_007(working: pd.DataFrame) -> pd.DataFrame:
    years = _fiscal_years(working)
    columns = ["Type", "Department", *_year_columns(years)]
    if working.empty or not years:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for load_type, label in [("load", "load"), ("contract", "contract"), ("all", "total")]:
        for department in _department_order(working):
            dept_df = working.loc[working["Department"].eq(department)].copy()
            if load_type != "all":
                dept_df = dept_df.loc[dept_df["load/contract"].eq(load_type)].copy()
            line = {"Type": label, "Department": department}
            for year in years:
                line[f"FY{year}"] = _safe_sum(dept_df.loc[dept_df["Fiscal Year"].eq(year), "Cr."])
            lines.append(line)
        total_line = {"Type": label, "Department": "total"}
        for year in years:
            subset = working.loc[working["Fiscal Year"].eq(year)].copy()
            if load_type != "all":
                subset = subset.loc[subset["load/contract"].eq(load_type)].copy()
            total_line[f"FY{year}"] = _safe_sum(subset["Cr."])
        lines.append(total_line)
        lines.append(_blank_row(columns))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), _year_columns(years))


def _budget_report_008a(working: pd.DataFrame) -> pd.DataFrame:
    years = _fiscal_years(working)
    columns = ["Program"]
    for year in years:
        columns.extend([f"FY{year} Credits", f"FY{year} Δ vs prev"])
    if working.empty or not years:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for program in _program_columns(working):
        program_df = working.loc[working["Program"].eq(program)].copy()
        line = {"Program": program}
        previous = 0.0
        for year in years:
            credits = _safe_sum(program_df.loc[program_df["Fiscal Year"].eq(year), "Cr."])
            line[f"FY{year} Credits"] = credits or pd.NA
            line[f"FY{year} Δ vs prev"] = (credits - previous) if years.index(year) > 0 or previous else pd.NA
            previous = credits
        lines.append(line)
    report = pd.DataFrame(lines, columns=columns)
    total_line = {"Program": "TOTAL"}
    for year in years:
        total_line[f"FY{year} Credits"] = _safe_sum(report[f"FY{year} Credits"])
    for idx, year in enumerate(years):
        total_line[f"FY{year} Δ vs prev"] = (
            total_line[f"FY{year} Credits"] - total_line[f"FY{years[idx - 1]} Credits"]
            if idx > 0
            else pd.NA
        )
    report = pd.concat([report, pd.DataFrame([total_line])], ignore_index=True)
    return _clean_numeric_display(report, [column for column in columns if column != "Program"])


def _budget_report_008b(working: pd.DataFrame) -> pd.DataFrame:
    years = _fiscal_years(working)
    columns = ["Department", "Program"]
    for year in years:
        columns.extend([f"FY{year} Credits", f"FY{year} Δ vs prev"])
    if working.empty or not years:
        return pd.DataFrame(columns=columns)

    lines: list[dict[str, object]] = []
    for department in _department_order(working):
        dept_df = working.loc[working["Department"].eq(department)].copy()
        if dept_df.empty:
            continue
        for program in _program_columns(dept_df):
            program_df = dept_df.loc[dept_df["Program"].eq(program)].copy()
            if program_df.empty:
                continue
            line = {"Department": department, "Program": program}
            previous = 0.0
            for idx, year in enumerate(years):
                credits = _safe_sum(program_df.loc[program_df["Fiscal Year"].eq(year), "Cr."])
                line[f"FY{year} Credits"] = credits or pd.NA
                line[f"FY{year} Δ vs prev"] = (credits - previous) if idx > 0 else pd.NA
                previous = credits
            lines.append(line)
        total_line = {"Department": department, "Program": "DEPARTMENT TOTAL"}
        previous = 0.0
        for idx, year in enumerate(years):
            credits = _safe_sum(dept_df.loc[dept_df["Fiscal Year"].eq(year), "Cr."])
            total_line[f"FY{year} Credits"] = credits or pd.NA
            total_line[f"FY{year} Δ vs prev"] = (credits - previous) if idx > 0 else pd.NA
            previous = credits
        lines.append(total_line)
        lines.append(_blank_row(columns))
        lines.append(_blank_row(columns))
    report = pd.DataFrame(lines, columns=columns)
    return _clean_numeric_display(report, [column for column in columns if column not in {"Department", "Program"}])


def _budget_report_009(working: pd.DataFrame) -> pd.DataFrame:
    years = _fiscal_years(working)
    max_cols = max(1 + len(years) * 2, 1 + len(years))
    columns = [f"col_{idx}" for idx in range(1, max_cols + 1)]
    if working.empty or not years:
        return pd.DataFrame(columns=columns)

    lines: list[list[object]] = []
    load_df = working.loc[working["load/contract"].eq("load")].copy()
    contract_df = working.loc[working["load/contract"].eq("contract")].copy()

    def _pad(row: list[object]) -> list[object]:
        return row + [pd.NA] * (len(columns) - len(row))

    def _simple_header(label: str) -> list[object]:
        return _pad([label, *[f"FY {year}" for year in years]])

    def _contract_header(label: str) -> list[object]:
        row = [label]
        for year in years:
            row.extend([f"FY {year}", f"% contracts to load FY {year}"])
        return _pad(row)

    def _simple_data_row(program: str, df: pd.DataFrame) -> list[object]:
        return _pad([program, *[_safe_sum(df.loc[df["Fiscal Year"].eq(year), "Cr."]) for year in years]])

    def _contract_data_row(program: str, df: pd.DataFrame) -> list[object]:
        row = [program]
        for year in years:
            contract_value = _safe_sum(df.loc[df["Fiscal Year"].eq(year), "Cr."])
            load_value = _safe_sum(
                load_df.loc[
                    load_df["Program"].eq(program) & load_df["Fiscal Year"].eq(year),
                    "Cr.",
                ]
            )
            ratio = contract_value / load_value if load_value else 0
            row.extend([contract_value, ratio])
        return _pad(row)

    programs = _program_columns(working)
    lines.append(_pad(["COMPARISON OF CREDITS BUDGETED per PROGRAM"]))
    lines.append(_pad([]))
    lines.append(_simple_header("TOTAL LOAD"))
    for program in programs:
        program_df = load_df.loc[load_df["Program"].eq(program)].copy()
        lines.append(_simple_data_row(program, program_df))
    lines.append(_simple_data_row("TOTAL", load_df))
    lines.extend([_pad([]), _pad([]), _pad([])])

    lines.append(_contract_header("TOTAL CONTRACT"))
    for program in programs:
        program_df = contract_df.loc[contract_df["Program"].eq(program)].copy()
        lines.append(_contract_data_row(program, program_df))
    total_contract_row = ["TOTAL"]
    for year in years:
        contract_value = _safe_sum(contract_df.loc[contract_df["Fiscal Year"].eq(year), "Cr."])
        load_value = _safe_sum(load_df.loc[load_df["Fiscal Year"].eq(year), "Cr."])
        total_contract_row.extend([contract_value, (contract_value / load_value) if load_value else 0])
    lines.append(_pad(total_contract_row))
    grand_total_row = ["GRAND TOTAL"]
    for year in years:
        grand_total_row.extend([_safe_sum(working.loc[working["Fiscal Year"].eq(year), "Cr."]), pd.NA])
    lines.append(_pad(grand_total_row))
    lines.extend([_pad([]), _pad([]), _pad([])])

    lines.append(_simple_header("TOTAL LOAD AND CONTRACT"))
    for program in programs:
        program_df = working.loc[working["Program"].eq(program)].copy()
        lines.append(_simple_data_row(program, program_df))
    lines.append(_simple_data_row("TOTAL", working))
    return _clean_numeric_display(pd.DataFrame(lines, columns=columns), columns)


def _budget_report_010(working: pd.DataFrame) -> pd.DataFrame:
    columns = ["Program", "NADEI_courses", "total_credits"]
    if working.empty:
        return pd.DataFrame(columns=columns)
    nadei = working.loc[_string_column(working, "Faculty Name").str.startswith("NADEI:", na=False)].copy()
    if nadei.empty:
        return pd.DataFrame(columns=columns)
    report = (
        nadei.groupby("Program", dropna=False)
        .agg(
            NADEI_courses=("Program", "size"),
            total_credits=("Cr.", "sum"),
        )
        .reset_index()
        .sort_values("Program", na_position="last")
    )
    return _clean_numeric_display(report, ["NADEI_courses", "total_credits"])


def _location_abbreviation(value: object) -> str:
    text = _safe_text(value)
    if not text:
        return ""
    if " - " in text:
        return text.rsplit(" - ", 1)[-1].strip()
    if text.upper() in {"MC", "ASYNC", "SYNC", "AU", "MC/DUAL", "DUAL"}:
        return text
    return text


def _course_display_for_contract(row: pd.Series) -> str:
    pieces = [_safe_text(row.get("Faculty Name"))]
    course = _safe_text(row.get("Course Text"))
    title = _safe_text(row.get("Catalog Title"))
    if course:
        pieces.append(course)
    if title:
        pieces.append(title)
    return " ".join(piece for piece in pieces if piece)


def _is_contract_value(value: object) -> bool:
    return _safe_text(value).lower() == "contract"


def _write_contract_block(ws, rows_df: pd.DataFrame, slot_rows: dict[str, list[int]], semesters: list[str]) -> None:
    sem_to_label = {semester: semester for semester in semesters}
    for semester, rows in slot_rows.items():
        label_row = rows[0] - 1
        ws.cell(label_row, 2).value = sem_to_label.get(semester, semester)
        semester_df = rows_df.loc[rows_df["Semester"].eq(semester)].copy()
        semester_df = semester_df.sort_values(["Faculty Name", "Course Text"], na_position="last")
        for target_row, (_, row) in zip(rows, semester_df.iterrows()):
            ws.cell(target_row, 2).value = _course_display_for_contract(row)
            ws.cell(target_row, 3).value = _location_abbreviation(row.get("Location"))
            ws.cell(target_row, 4).value = row.get("Cr.")
            ws.cell(target_row, 5).value = row.get("rate per credit")
            ws.cell(target_row, 7).value = f"=+F{target_row}"
            ws.cell(target_row, 8).value = None


def generate_department_contracts_overview_report(df: pd.DataFrame, department: str) -> ReportResult:
    working = _filter_department(_prep_working_df(df), department)
    top_semesters = _department_cycle_semesters(working["Semester"]) if not working.empty else []
    start_year = _academic_year_sort_key(top_semesters[0])[0] if top_semesters else None
    workbook = _build_contracts_overview_template(start_year)
    ws = workbook.active

    if top_semesters:
        ws["B4"] = f"May {start_year}-April {start_year + 1}"
        ws["B26"] = f"May {start_year} - April {start_year + 1}"
    semester_rows = {7: f"Summer {start_year}" if start_year is not None else "", 8: f"Fall {start_year}" if start_year is not None else "", 9: f"Spring {start_year + 1}" if start_year is not None else ""}
    for row_idx, semester in semester_rows.items():
        ws.cell(row_idx, 2).value = f"Credits on {semester} schedule" if semester else None
        ws.cell(row_idx, 6).value = _safe_sum(working.loc[working["Semester"].eq(semester), "Cr."]) if semester else None

    load_df = working.loc[working["load/contract"].eq("load")].copy()
    salaried = (
        load_df.groupby("Faculty Name", dropna=False)["Cr."].sum().reset_index().sort_values(["Cr.", "Faculty Name"], ascending=[False, True])
        if not load_df.empty
        else pd.DataFrame(columns=["Faculty Name", "Cr."])
    )
    for row_idx in range(14, 19):
        ws.cell(row_idx, 2).value = None
        ws.cell(row_idx, 6).value = None
    for row_idx, (_, row) in zip(range(14, 19), salaried.iterrows()):
        ws.cell(row_idx, 2).value = row["Faculty Name"]
        ws.cell(row_idx, 6).value = row["Cr."]

    contract_df = working.loc[working["load/contract"].eq("contract")].copy()
    extension_programs = {
        "MAPM-ENGL": "mapm_engl",
        "MAPM-HISP": "mapm_hisp",
        "MAR-INT": "ma_religion",
        "MDivHISP_MAPmENGL_MAPmHISP": "mdiv_spanish",
    }
    extension_mask = contract_df["Program"].isin(extension_programs)
    on_campus_df = contract_df.loc[~extension_mask].copy()
    start_year = start_year or 0
    on_campus_slots = {
        f"Summer {start_year}": [35, 36, 37],
        f"Fall {start_year}": [40, 41, 42, 43, 44, 45],
        f"Spring {start_year + 1}": [48, 49, 50, 51, 52, 53],
    }
    _write_contract_block(ws, on_campus_df, on_campus_slots, top_semesters)

    extension_sections = {
        "MAPM-ENGL": {f"Fall {start_year}": [60, 61], f"Spring {start_year + 1}": [63, 64, 65, 66]},
        "MAPM-HISP": {f"Summer {start_year}": [69, 70], f"Fall {start_year}": [72, 73, 74], f"Spring {start_year + 1}": [76, 77]},
        "MAR-INT": {f"Fall {start_year}": [81, 82, 83, 84]},
        "MDivHISP_MAPmENGL_MAPmHISP": {f"Fall {start_year}": [88, 89, 90, 91, 92]},
    }
    for program, slot_rows in extension_sections.items():
        program_df = contract_df.loc[contract_df["Program"].eq(program)].copy()
        _write_contract_block(ws, program_df, slot_rows, top_semesters)

    output = BytesIO()
    workbook.save(output)
    preview_df = contract_df.loc[
        :,
        [column for column in ["Semester", "Faculty Name", "Course Text", "Program", "Location", "Cr.", "rate per credit", "total contract amount"] if column in contract_df.columns],
    ].copy()
    preview_df = preview_df.rename(columns={"Course Text": "Course"})
    return ReportResult(
        report_name=f"{department} Contracts Overview",
        preview_df=preview_df,
        excel_bytes=output.getvalue(),
        output_filename=f"{department}_ContractsOverview.xlsx",
        worksheet_name="Contracts",
        metadata={"department": department},
    )


def generate_department_overview_bundle(df: pd.DataFrame, department: str) -> ReportResult:
    working = _filter_department(_prep_working_df(df), department)
    sheets = {
        "report_001a": _overview_report_001a(working),
        "report_001b": _overview_report_001b(working),
        "report_001c": _overview_report_001c(working),
        "report_001f": _overview_report_001f(working),
        "report_002a": _overview_report_002a(working),
        "report_003a": _overview_report_003a(working),
    }
    workbook_bytes = workbook_to_excel_bytes(sheets, style_plan={key: key for key in sheets})
    contracts_report = generate_department_contracts_overview_report(df, department)
    workbook = load_workbook(BytesIO(workbook_bytes))
    contracts_workbook = load_workbook(BytesIO(contracts_report.excel_bytes))
    _clone_sheet_into_workbook(contracts_workbook.active, workbook, "ContractsOverview")
    combined_buffer = BytesIO()
    workbook.save(combined_buffer)
    preview_df = sheets["report_001a"]
    return ReportResult(
        report_name=f"{department} Department Overview Reports",
        preview_df=preview_df,
        excel_bytes=combined_buffer.getvalue(),
        output_filename=f"{department}_department_overview_reports_demo.xlsx",
        worksheet_name="report_001a",
        metadata={"department": department, "sheet_count": str(len(sheets) + 1)},
    )


def generate_contract_forms_demo(df: pd.DataFrame) -> ReportResult:
    working = _prep_working_df(df)
    if not working.empty:
        working = working.loc[working["load/contract"].eq("contract")].copy()
    selected_columns = [
        "Department",
        "Semester",
        "Faculty Name",
        "ID#",
        "email",
        "Course",
        "Catalog Title",
        "Cr.",
        "rate per credit",
        "total contract amount",
        "Begin Date",
        "End Date",
        "account to be charged",
        "Location",
        "Mode",
        "Source Workbook",
        "Source Sheet",
    ]
    preview_df = working.loc[:, [c for c in selected_columns if c in working.columns]].copy()
    preview_df = preview_df.rename(columns={"Cr.": "Credits"})
    sheets = {
        "contract_rows": preview_df,
        "summary_by_faculty": (
            preview_df.groupby(["Faculty Name", "Semester"], dropna=False)
            .agg(
                Credits=("Credits", "sum"),
                Contract_Cost=("total contract amount", "sum"),
                Course_Count=("Course", lambda s: s.dropna().nunique()),
            )
            .reset_index()
            .rename(columns={"Contract_Cost": "Contract Cost", "Course_Count": "Course Count"})
            .sort_values(["Faculty Name", "Semester"], na_position="last")
            if not preview_df.empty
            else pd.DataFrame(columns=["Faculty Name", "Semester", "Credits", "Contract Cost", "Course Count"])
        ),
    }
    return ReportResult(
        report_name="Contract Forms Demo Export",
        preview_df=preview_df,
        excel_bytes=workbook_to_excel_bytes(sheets),
        output_filename="contract_forms_demo_from_donsheet.xlsx",
        worksheet_name="contract_rows",
        metadata={"rows": str(len(preview_df))},
    )


def generate_schedule_demo(df: pd.DataFrame) -> ReportResult:
    working = _prep_working_df(df)
    day_columns = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]
    for column in day_columns:
        working[column] = _string_column(working, column)
    schedule_columns = [
        "Department",
        "Semester",
        "Course",
        "Catalog Title",
        "Faculty Name",
        "Mode",
        "Location",
        "Course Beginning Time {Meet Beg Time}",
        "Course Ending Time {Meet End Time}",
        *day_columns,
        "Room {Meet Room}",
        "Building {Meet Bldg}",
        "Source Sheet",
    ]
    preview_df = working.loc[:, [c for c in schedule_columns if c in working.columns]].copy()
    preview_df = preview_df.rename(
        columns={
            "Course Beginning Time {Meet Beg Time}": "Begin Time",
            "Course Ending Time {Meet End Time}": "End Time",
            "Room {Meet Room}": "Room",
            "Building {Meet Bldg}": "Building",
        }
    )
    preview_df = preview_df.sort_values(["Semester", "Department", "Course"], na_position="last")
    summary_df = (
        preview_df.groupby(["Semester", "Department"], dropna=False)
        .agg(
            Course_Count=("Course", lambda s: s.dropna().nunique()),
            Faculty_Count=("Faculty Name", lambda s: s.dropna().nunique()),
        )
        .reset_index()
        .rename(columns={"Course_Count": "Course Count", "Faculty_Count": "Faculty Count"})
        if not preview_df.empty
        else pd.DataFrame(columns=["Semester", "Department", "Course Count", "Faculty Count"])
    )
    return ReportResult(
        report_name="SmartSchedule Demo Export",
        preview_df=preview_df,
        excel_bytes=workbook_to_excel_bytes({"schedule_rows": preview_df, "schedule_summary": summary_df}),
        output_filename="smart_schedule_demo_from_donsheet.xlsx",
        worksheet_name="schedule_rows",
        metadata={"rows": str(len(preview_df))},
    )


def _budget_report_master(working: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Source Workbook",
        "Department",
        "Semester",
        "Fiscal Year",
        "Faculty Name",
        "Course",
        "Catalog Title",
        "Program",
        "Mode",
        "Location",
        "load/contract",
        "Cr.",
        "rate per credit",
        "total contract amount",
        "account to be charged",
    ]
    available = [column for column in columns if column in working.columns]
    report = working.loc[:, available].copy()
    return report.sort_values(["Department", "Semester", "Faculty Name", "Course"], na_position="last")


def generate_budget_reports_bundle(df: pd.DataFrame, source_label: str = "donsheet") -> ReportBundle:
    working = _prep_working_df(df)
    sheet_frames = {
        "report_000": _budget_report_master(working),
        "report_001a": _overview_report_001a(working),
        "report_001b": _overview_report_001b(working),
        "report_001c": _overview_report_001c(working),
        "report_001f": _overview_report_001f_all_departments(working),
        "report_002a": _overview_report_002a(working),
        "report_002b": _budget_report_002b(working),
        "report_003a": _overview_report_003a(working),
        "report_003b": _budget_report_003b(working),
        "report_004a": _budget_report_004a(working),
        "report_004b": _budget_report_004b(working),
        "report_005": _budget_report_005(working),
        "report_006": _budget_report_006(working),
        "report_007": _budget_report_007(working),
        "report_008a": _budget_report_008a(working),
        "report_008b": _budget_report_008b(working),
        "report_009": _budget_report_009(working),
        "report_010": _budget_report_010(working),
    }
    report_titles = {
        "report_000": "Master Data",
        "report_001a": "Department Overview",
        "report_001b": "Department Contract Overview",
        "report_001c": "Department Overview by Load Type",
        "report_001f": "Department Faculty Year Overview",
        "report_002a": "Department Credit Cost Overview",
        "report_002b": "Department Courses by Program Credits",
        "report_003a": "Department Totals Load Contract",
        "report_003b": "Subject Totals Load Contract",
        "report_004a": "Program Credits by Department",
        "report_004b": "Program Load Credits by Department",
        "report_005": "Program Totals Load Contract",
        "report_006": "Program Contract Cost per Department",
        "report_007": "FY Comparison Department Totals",
        "report_008a": "Program Credits Year Pairs",
        "report_008b": "FY Comparison Department Program Credit",
        "report_009": "FY Comparison Program Load Credits",
        "report_010": "Teaching per Program",
    }
    reports: dict[str, ReportResult] = {}
    for key, frame in sheet_frames.items():
        reports[key] = ReportResult(
            report_name=report_titles[key],
            preview_df=frame,
            excel_bytes=workbook_to_excel_bytes({key: frame}, style_plan={key: key}),
            output_filename=f"{key}_from_{source_label}.xlsx",
            worksheet_name=key,
            metadata={"rows": str(len(frame))},
        )

    combined_frames = {**sheet_frames, "report_999": _budget_report_master(working)}
    return ReportBundle(
        bundle_name="Budget Reports",
        reports=reports,
        combined_workbook_bytes=workbook_to_excel_bytes(
            combined_frames,
            style_plan={key: key for key in combined_frames},
        ),
        combined_output_filename=f"report_999_budget_reports_from_{source_label}.xlsx",
        preview_report_key="report_001a",
        metadata={"sheet_count": str(len(combined_frames))},
    )
