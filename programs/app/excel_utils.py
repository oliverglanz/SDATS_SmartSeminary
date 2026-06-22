from __future__ import annotations

from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def dataframe_to_excel_bytes(dataframe: pd.DataFrame, sheet_name: str) -> bytes:
    return workbook_to_excel_bytes({sheet_name: dataframe})


def workbook_to_excel_bytes(
    sheets: dict[str, pd.DataFrame],
    style_plan: dict[str, str] | None = None,
) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_map: dict[str, tuple[str, pd.DataFrame]] = {}
        for idx, (sheet_name, dataframe) in enumerate(sheets.items(), start=1):
            safe_sheet_name = _safe_sheet_name(sheet_name, idx)
            style_key = (style_plan or {}).get(sheet_name, sheet_name)
            write_header = style_key != "report_009"
            dataframe.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=write_header)
            sheet_map[safe_sheet_name] = (sheet_name, dataframe)
        for safe_sheet_name, (original_name, dataframe) in sheet_map.items():
            worksheet = writer.book[safe_sheet_name]
            style_key = (style_plan or {}).get(original_name, original_name)
            _apply_sheet_style(worksheet, dataframe, style_key)
    buffer.seek(0)
    return buffer.read()


def _safe_sheet_name(sheet_name: str, fallback_index: int) -> str:
    cleaned = str(sheet_name).strip() or f"Sheet{fallback_index}"
    cleaned = cleaned.replace("/", "-").replace("\\", "-").replace("*", "")
    cleaned = cleaned.replace("[", "(").replace("]", ")").replace(":", " -")
    cleaned = cleaned.replace("?", "").strip()
    return cleaned[:31] or f"Sheet{fallback_index}"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
ALT_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
SUMMER_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FABF8F")
SUMMER_DETAIL_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
FALL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="C4D79B")
FALL_DETAIL_FILL = PatternFill(fill_type="solid", fgColor="EBF1DE")
SPRING_HEADER_FILL = PatternFill(fill_type="solid", fgColor="8DB4E2")
SPRING_DETAIL_FILL = PatternFill(fill_type="solid", fgColor="C5D9F1")
HEADER_FONT = Font(bold=True, name="Arial", size=11)
BODY_FONT = Font(name="Arial", size=11)
TITLE_FONT = Font(bold=True, color="0000FF", size=14, name="Arial")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
TOP_LEFT = Alignment(horizontal="left", vertical="top")
TOP_CENTER = Alignment(horizontal="center", vertical="top")
BLACK_THIN = Side(style="thin", color="000000")
BLACK_MEDIUM = Side(style="medium", color="000000")


def _apply_sheet_style(ws, dataframe: pd.DataFrame, style_key: str) -> None:
    if ws.max_row == 0 or ws.max_column == 0:
        return
    _apply_base_font(ws)
    if style_key == "report_001a":
        _style_flat_report(
            ws,
            currency_headers={"Cost", "Check Total", *dataframe.columns[9:-2]},
            decimal_currency=False,
            striped=True,
            summary_markers={"DEPARTMENT TOTAL"},
        )
    elif style_key == "report_001b":
        _style_sectioned_contract_report(ws, credits_header="Credits", rate_header="Rate", total_header="Total")
    elif style_key == "report_001c":
        _style_sectioned_semester_report(ws, credits_header="Credits", rate_header="Rate", total_header="Total")
    elif style_key == "report_001f":
        _style_report_001f(ws)
    elif style_key == "report_002a":
        _style_credit_cost_report(ws)
        _highlight_keyword_rows(ws, {"DEPARTMENT TOTAL"})
    elif style_key == "report_002b":
        _style_generic_table(ws)
        _format_columns_by_header_keywords(ws)
        _highlight_keyword_rows(ws, {"DEPARTMENT TOTAL"})
    elif style_key == "report_003a":
        _style_totals_report(ws)
    elif style_key == "report_003b":
        _style_totals_report(ws)
        _highlight_keyword_rows(ws, {"TOTAL"})
    elif style_key in {"report_004a", "report_004b", "report_006", "report_008b"}:
        _style_generic_table(ws)
        _format_columns_by_header_keywords(ws)
        _highlight_keyword_rows(ws, {"DEPARTMENT TOTAL", "DEPARTMENT TOTAL LOAD CREDITS", "DEPARTMENT TOTAL CONTRACT CREDITS", "DEPARTMENT TOTAL CONTRACT+LOAD CREDITS"})
    elif style_key == "report_005":
        _style_generic_table(ws)
        _format_columns_by_header_keywords(ws)
        _highlight_keyword_rows(ws, {"TOTAL"})
    elif style_key in {"report_007", "report_009"}:
        if style_key == "report_007":
            _style_generic_table(ws)
            _format_columns_by_header_keywords(ws)
            _highlight_keyword_rows(ws, {"TOTAL"})
        else:
            _style_report_009(ws)
    elif style_key == "report_008a":
        _style_generic_table(ws)
        _format_columns_by_header_keywords(ws)
        _highlight_keyword_rows(ws, {"TOTAL"})
    elif style_key == "report_010":
        _style_generic_table(ws)
        _format_columns_by_header_keywords(ws)
    else:
        _style_generic_table(ws)


def _apply_base_font(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            cell.font = Font(
                name="Arial",
                size=11,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
            )


def _style_generic_table(ws) -> None:
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _style_flat_report(
    ws,
    currency_headers: set[str],
    decimal_currency: bool,
    striped: bool,
    summary_markers: set[str] | None = None,
) -> None:
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if striped:
        for row_idx in range(2, ws.max_row + 1):
            fill = ALT_FILL if row_idx % 2 == 1 else WHITE_FILL
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "").strip()
        if header in currency_headers:
            _format_column(ws, col_idx, '$#,##0.00' if decimal_currency else '$#,##0')
        elif header in {"Cr.", "Credits"} or "Credits" in header:
            _format_column(ws, col_idx, '0.##')
    if summary_markers:
        _highlight_keyword_rows(ws, summary_markers)
    _autosize_columns(ws)


def _style_sectioned_contract_report(ws, credits_header: str, rate_header: str, total_header: str) -> None:
    _style_plain_header_row(ws, 1)
    _format_named_column(ws, credits_header, "0.##")
    _format_named_column(ws, rate_header, '"$"#,##0.00')
    _format_named_column(ws, total_header, '"$"#,##0.00')
    for row_idx in range(2, ws.max_row + 1):
        first = str(ws.cell(row_idx, 1).value or "").strip()
        rest_empty = all(ws.cell(row_idx, col).value in (None, "") for col in range(2, ws.max_column + 1))
        if _is_department_title(first, rest_empty):
            _style_title_row(ws, row_idx)
        elif first.lower().endswith(" contracts"):
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
        elif first.startswith("Total "):
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
        elif first and rest_empty:
            _font_row(ws, row_idx, HEADER_FONT)
    _autosize_columns(ws)


def _style_sectioned_semester_report(ws, credits_header: str, rate_header: str, total_header: str) -> None:
    _style_plain_header_row(ws, 1)
    _format_named_column(ws, credits_header, "0.##")
    _format_named_column(ws, rate_header, '"$"#,##0.00')
    _format_named_column(ws, total_header, '"$"#,##0.00')
    for row_idx in range(2, ws.max_row + 1):
        first = str(ws.cell(row_idx, 1).value or "").strip()
        rest_empty = all(ws.cell(row_idx, col).value in (None, "") for col in range(2, ws.max_column + 1))
        if _is_department_title(first, rest_empty):
            _style_title_row(ws, row_idx)
        elif first.startswith("Total "):
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
        elif first and rest_empty:
            _font_row(ws, row_idx, HEADER_FONT)
    _autosize_columns(ws)


def _style_credit_cost_report(ws) -> None:
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "").strip()
        if header == "Cost" or header.endswith(" Cost"):
            _format_column(ws, col_idx, '"$"#,##0.00')
        elif header in {"Credits from Load", "Credits from Contract", "Cr."} or header.endswith(" Credits"):
            _format_column(ws, col_idx, "0.##")
    _autosize_columns(ws)


def _style_totals_report(ws) -> None:
    _style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "").strip()
        if header == "Contract Cost":
            _format_column(ws, col_idx, "$#,##0")
        elif "Ratio" in header:
            _format_column(ws, col_idx, "0.00")
        elif "credits" in header.lower() or "Credits" in header:
            _format_column(ws, col_idx, "0.##")
    _autosize_columns(ws)


def _style_header_row(ws, row_idx: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _style_plain_header_row(ws, row_idx: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _style_title_row(ws, row_idx: int) -> None:
    cell = ws.cell(row=row_idx, column=1)
    cell.font = TITLE_FONT
    cell.alignment = LEFT


def _style_report_009(ws) -> None:
    title = ws.cell(1, 1)
    title.font = Font(bold=True, color="FF0000", size=14, name="Arial")
    title.alignment = LEFT

    active_section = None
    for row_idx in range(1, ws.max_row + 1):
        first = str(ws.cell(row_idx, 1).value or "").strip()
        if first in {"TOTAL LOAD", "TOTAL CONTRACT", "TOTAL LOAD AND CONTRACT"}:
            active_section = first
            _fill_row(ws, row_idx, HEADER_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = CENTER
        elif first in {"TOTAL", "GRAND TOTAL"}:
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
        elif first:
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell.value, (int, float)):
                    continue
                if active_section == "TOTAL CONTRACT" and col_idx >= 3 and col_idx % 2 == 1:
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "0.##"
                cell.alignment = RIGHT
    _autosize_columns(ws)


def _style_report_001f(ws) -> None:
    _style_plain_header_row(ws, 1)
    current_faculty_start = None
    current_department_start = None

    def _semester_fills(course_text: str):
        lowered = course_text.lower()
        if "summer" in lowered:
            return SUMMER_HEADER_FILL, SUMMER_DETAIL_FILL
        if "fall" in lowered:
            return FALL_HEADER_FILL, FALL_DETAIL_FILL
        if "spring" in lowered:
            return SPRING_HEADER_FILL, SPRING_DETAIL_FILL
        return None, None

    active_detail_fill = None

    for row_idx in range(2, ws.max_row + 1):
        department = ws.cell(row_idx, 1).value
        faculty = ws.cell(row_idx, 2).value
        course = str(ws.cell(row_idx, 3).value or "").strip()

        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).alignment = TOP_LEFT

        if faculty:
            if current_faculty_start is not None and row_idx - 1 > current_faculty_start:
                ws.merge_cells(start_row=current_faculty_start, start_column=2, end_row=row_idx - 1, end_column=2)
            current_faculty_start = row_idx
            ws.cell(row_idx, 2).fill = HEADER_FILL
            ws.cell(row_idx, 2).font = HEADER_FONT
            ws.cell(row_idx, 2).alignment = TOP_CENTER
            _apply_top_border(ws, row_idx, BLACK_THIN)
        if department:
            if current_department_start is not None and row_idx - 1 > current_department_start:
                ws.merge_cells(start_row=current_department_start, start_column=1, end_row=row_idx - 1, end_column=1)
            current_department_start = row_idx
            cell = ws.cell(row_idx, 1)
            cell.font = TITLE_FONT
            cell.alignment = TOP_CENTER
            _apply_top_border(ws, row_idx, BLACK_MEDIUM)
        if course.startswith("Total "):
            _font_row(ws, row_idx, HEADER_FONT)
            header_fill, _ = _semester_fills(course)
            if header_fill:
                for col_idx in range(3, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = header_fill
        if course.endswith("2026") or course.endswith("2027") or course.endswith("2028"):
            ws.cell(row_idx, 3).font = HEADER_FONT
            header_fill, detail_fill = _semester_fills(course)
            active_detail_fill = detail_fill
            if header_fill:
                for col_idx in range(3, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = header_fill
        if course == "TOTAL":
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)
            for col_idx in range(3, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = SECTION_FILL
            _apply_bottom_border(ws, row_idx, BLACK_THIN)
        elif course and not any(keyword in course for keyword in ["Total ", "TOTAL"]) and ws.cell(row_idx, 4).value is None:
            ws.cell(row_idx, 3).font = HEADER_FONT
        elif course and ws.cell(row_idx, 4).value is not None and active_detail_fill:
            for col_idx in range(3, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = active_detail_fill

    if current_faculty_start is not None and ws.max_row > current_faculty_start:
        ws.merge_cells(start_row=current_faculty_start, start_column=2, end_row=ws.max_row, end_column=2)
    if current_department_start is not None and ws.max_row >= current_department_start:
        ws.merge_cells(start_row=current_department_start, start_column=1, end_row=ws.max_row, end_column=1)

    for col_idx in range(8, 11):
        if col_idx <= ws.max_column:
            fmt = '"$"#,##0.00' if col_idx == 10 else "0.##"
            _format_column(ws, col_idx, fmt)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.alignment.horizontal == "right":
                cell.alignment = Alignment(horizontal="right", vertical="top")
    _autosize_columns(ws)


def _is_department_title(first: str, rest_empty: bool) -> bool:
    if not first or not rest_empty:
        return False
    if any(char.isdigit() for char in first):
        return False
    if first.lower().endswith(" contracts") or first.startswith("Total "):
        return False
    return True


def _fill_row(ws, row_idx: int, fill) -> None:
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _apply_top_border(ws, row_idx: int, side: Side) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = Border(
            left=copy(cell.border.left),
            right=copy(cell.border.right),
            top=side,
            bottom=copy(cell.border.bottom),
        )


def _apply_bottom_border(ws, row_idx: int, side: Side) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = Border(
            left=copy(cell.border.left),
            right=copy(cell.border.right),
            top=copy(cell.border.top),
            bottom=side,
        )


def _font_row(ws, row_idx: int, font) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(
            name=font.name,
            size=font.size,
            bold=font.bold,
            italic=font.italic,
            color=font.color.rgb if font.color and font.color.type == "rgb" else None,
        )


def _format_named_column(ws, header_name: str, fmt: str) -> None:
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(2, col_idx).value or "").strip() == header_name or str(ws.cell(1, col_idx).value or "").strip() == header_name:
            _format_column(ws, col_idx, fmt)


def _format_column(ws, col_idx: int, fmt: str) -> None:
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt
            cell.alignment = RIGHT


def _format_columns_by_header_keywords(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "").strip()
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ["cost", "amount", "dollars", "rate"]):
            _format_column(ws, col_idx, '"$"#,##0.00')
        elif any(keyword in header_lower for keyword in ["credits", "credit", "count", "ratio", "fy20", "% contract-to-load"]):
            _format_column(
                ws,
                col_idx,
                "0.00" if ("ratio" in header_lower or "% contract-to-load" in header_lower) else "0.##",
            )


def _highlight_keyword_rows(ws, markers: set[str]) -> None:
    markers_upper = {marker.upper() for marker in markers}
    for row_idx in range(2, ws.max_row + 1):
        texts = [
            str(ws.cell(row_idx, col_idx).value or "").strip().upper()
            for col_idx in range(1, ws.max_column + 1)
        ]
        if any(text in markers_upper for text in texts):
            _fill_row(ws, row_idx, SECTION_FILL)
            _font_row(ws, row_idx, HEADER_FONT)


def _autosize_columns(ws, min_width: int = 8, max_width: int = 60, padding: int = 2) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value).replace("\n", " ")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + padding, min_width), max_width)
