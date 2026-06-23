from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, GradientFill, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


OUTPUT_FILENAME = "ColorReport_SEM_course_updates.xlsx"
BASE_KEY = "CRN"
HELPER_MATCH_COLUMNS = ["Instructor Email {Instr Email}"]
HELPER_SHEETS = {"DropDownMenu", "InstructorMap"}

BANNER_RENAMES = {
    "Crse Num": "Course Number {Crse Num}",
    "Seq Crse Num": "Course Section {Seq Crse Num}",
    "Sect Crs": "Credits {Sect Crs}",
    "Sect Sch Type": "Sect Sch Type",
    "Inst Method": "Instruction Method {Inst Method}",
    "Level SC": "Level {Level SC}",
    "Max Enrl": "Enrollment Cap {Max Enrl}",
    "Meet Start Date": "Course Start Date {Meet Start Date}",
    "Meet End Date": "Course End Date {Meet End Date}",
    "Meet Beg Time": "Course Beginning Time {Meet Beg Time}",
    "Meet End Time": "Course Ending Time {Meet End Time}",
    "Meet Bldg": "Building {Meet Bldg}",
    "Meet Room": "Room {Meet Room}",
    "Instr Name": "Instructor Name {Instr Name}",
    "Instr Email": "Instructor Email {Instr Email}",
    "Instr ID": "Instructor ID {Instr ID}",
    "Cat Crs": "Credit Catalog {Cat Crs}",
    "Preq Areas": "Prerequisites {Preq Areas}",
    "Camp Restr CC": "Campus Restriction {Camp Restr CC}",
    "Scacrse Dept": "SEM Department {Scacrse Dept}",
    "Special Apvl": "Special Approval {Special Apvl}",
    "Ssasect Campus": "Campus {Ssasect Campus}",
    "Soaterm Start Date": "Semester Start Date {Soaterm Start Date}",
    "Soaterm End Date": "Semester End Date {Soaterm End Date}",
    "Meet Override": "Course Override {Meet Override}",
    "Fees Amt": "Fee Amount {Fees Amt}",
    "Detl Code": "Fee Code {Detl Code}",
}

COMPARISON_COLUMNS = {
    "Subject": "Subject",
    "Course Number {Crse Num}": "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}": "Course Section {Seq Crse Num}",
    "Program {not in Banner}": "Program {not in Banner}",
    "Catalog Title": "Catalog Title",
    "Section Title": "Section Title",
    "Schedule Type {not in Banner}": "Sect Sch Type",
    "Instruction Method {Inst Method}": "Instruction Method {Inst Method}",
    "Level {Level SC}": "Level {Level SC}",
    "Credits {Sect Crs}": "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}": "Enrollment Cap {Max Enrl}",
    "Pre-work Start Date {not in Banner}": "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}": "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}": "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}": "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}": "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}": "Post-work End Date {not in Banner}",
    "Meeting Type": "Meeting Type",
    "Course Beginning Time {Meet Beg Time}": "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}": "Course Ending Time {Meet End Time}",
    "SUN": "SUN", "MON": "MON", "TUE": "TUE", "WED": "WED", "THU": "THU", "FRI": "FRI",
    "Room {Meet Room}": "Room {Meet Room}",
    "Building {Meet Bldg}": "Building {Meet Bldg}",
    "Instructor Name {Instr Name}": "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}": "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}": "Instructor ID {Instr ID}",
    "% Responsibility": "% Responsibility",
    "load/contract {not in Banner}": "load/contract {not in Banner}",
    "costs per credit {not in Banner}": "costs per credit {not in Banner}",
    "total costs {not in Banner}": "total costs {not in Banner}",
    "account to be charged {not in Banner}": "account to be charged {not in Banner}",
    "Weeks": "Weeks", "Term Code": "Term Code", "Part-Of-Term": "Part-Of-Term", "Camp SC": "Camp SC",
}

BANNER_RELEVANT_COLUMNS = [
    "CRN", "Subject", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "Section Title",
    "Instruction Method {Inst Method}", "Credits {Sect Crs}", "Enrollment Cap {Max Enrl}", "Meeting Type",
    "Course Start Date {Meet Start Date}", "Course End Date {Meet End Date}",
    "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}",
    "SUN", "MON", "TUE", "WED", "THU", "FRI", "Room {Meet Room}", "Building {Meet Bldg}",
    "Instructor Name {Instr Name}", "Instructor Email {Instr Email}", "Instructor ID {Instr ID}", "% Responsibility",
]

# The notebook uses this same gradient for every discrepancy category.
DISCREPANCY_FILL = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)
SUMMARY_HEADER_DEFAULT_FILL = PatternFill("solid", fgColor="FFD3D3D3")
SUMMARY_HEADER_BLUE_FILL = PatternFill("solid", fgColor="FFDDEBF7")
SUMMARY_HEADER_GREY_FILL = PatternFill("solid", fgColor="FFD9D9D9")

SITE_NAMES = {
    "ACW": "Wellness Center [ACW]", "BH": "Bell Hall [BH]", "BUL": "Buller Hall [BUL]",
    "CSH": "Chan Shun Hall [CSH]", "GHA": "Griggs Hall A [GHA]", "GHB": "Griggs Hall B [GHB]",
    "HORN": "Horn Museum [HORN]", "JGYM": "Johnson Gym [JGYM]", "JWL": "James White Library [JWL]",
    "NH": "Nethery Hall [NH]", "OCARG": "Univ Adventista del Plata [OCARG]",
    "OCBRA1": "Centro Univ Adv de Sao Paulo [OCBRA1]", "OCCABU": "Burman University [OCCABU]",
    "OCCALL": "Loma Linda University [OCCALL]", "OCCASC": "Southeastern California Conf [OCCASC]",
    "OCCHK": "Hong Kong Adventist College [OCCHK]", "OCFLAH": "Advent Health University [OCFLAH]",
    "OCFLFC": "Florida Conference of SDA [OCFLFC]", "OCFLFL": "Forest Lake SDA Church [OCFLFL]",
    "OCGBR": "Newbold College [OCGBR]", "OCMDND": "North American Division of SDA [OCMDND]",
    "OCMENE": "Northern New England Conf [OCMENE]", "OCNEUC": "Union College [OCNEUC]",
    "OCOKOC": "Oklahoma City Central [OCOKOC]", "OCPOL": "Polish Senior College Theo&Hum [OCPOL]",
    "OCROU": "Universitatea Adventus din Cer [OCROU]", "OCRUS": "Zaokski Theo Seminary [OCRUS]",
    "OCTHA": "Asia-Pacific Int Univ AIU [OCTHA]", "OCTWN": "Taiwan Adventist College [OCTWN]",
    "OCUKR": "Ukrainan Adv Center Higher ED [OCUKR]", "OCWANP": "North Pacific Union Conference [OCWANP]",
    "PMC": "Pioneer Memorial Church [PMC]", "SEM": "Seminary Building [SEM]",
}

DON_ONLY_BANNER_COLUMNS = [
    "Don {not in Banner}", "Mismatching {not in Banner}", "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}", "Karen {not in Banner}", "DONE {not in Banner}", "Notes {not in Banner}",
    "Program {not in Banner}", "Crosslist Details {not in Banner}", "Schedule Type {not in Banner}",
    "Pre-work Start Date {not in Banner}", "Pre-work End Date {not in Banner}",
    "Post-work Start Date {not in Banner}", "Post-work End Date {not in Banner}",
    "load/contract {not in Banner}", "costs per credit {not in Banner}", "total costs {not in Banner}",
    "account to be charged {not in Banner}",
]

BANNER_OUTPUT_COLUMNS = [
    "Don {not in Banner}", "Mismatching {not in Banner}", "Dept/Prog Admin {not in Banner}", "Mona {not in Banner}", "Karen {not in Banner}", "DONE {not in Banner}", "Notes {not in Banner}",
    "CRN sorted {not in Banner}", "CRN", "Subject", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "Program {not in Banner}", "Catalog Title", "Section Title", "Crosslist Details {not in Banner}", "X Lst", "Campus {Ssasect Campus}", "Campus Restriction {Camp Restr CC}", "Schedule Type {not in Banner}", "Instruction Method {Inst Method}", "Level {Level SC}", "Credit Catalog {Cat Crs}", "Credits {Sect Crs}", "Enrollment Cap {Max Enrl}", "Meeting Type", "Semester Start Date {Soaterm Start Date}", "Pre-work Start Date {not in Banner}", "Pre-work End Date {not in Banner}", "Course Start Date {Meet Start Date}", "Course End Date {Meet End Date}", "Post-work Start Date {not in Banner}", "Post-work End Date {not in Banner}", "Semester End Date {Soaterm End Date}", "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}", "SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT", "Room {Meet Room}", "Building {Meet Bldg}", "Instructor Name {Instr Name}", "Instructor Email {Instr Email}", "Instructor ID {Instr ID}", "% Responsibility", "load/contract {not in Banner}", "costs per credit {not in Banner}", "total costs {not in Banner}", "account to be charged {not in Banner}",
    "Fee Amount {Fees Amt}", "Fee Code {Detl Code}", "Fee Type", "Fee Level", "Fee Ind", "Fee Term", "SEM Department {Scacrse Dept}", "Activity Date", "Status", "Bill Hrs", "Sect Sch Type", "Prerequisites {Preq Areas}", "Link Ind", "Integration Code", "Link Conn", "Crs Attr", "Comments", "Major Restr", "Scacrse College", "Enrolled", "Waitlist Capacity", "Primary Ind", "Override", "1st Date Reg Opens", "Last Date Reg Opens", "OL Range From Date", "OL Range To Date", "OL Numb Units", "OL Duration Code", "Grade Mode", "Gradable Ind", "Tuit Waiver", "Course Override {Meet Override}", "Lvl Res Ind", "Cmp Res Ind", "Rate Code", "Cohort Code", "St Attr Code", "Coll SC", "DegC SC", "Prog SC", "FOS Type", "Special Approval {Special Apvl}", "Weeks", "Term Code", "Part-Of-Term", "Camp SC",
    "_RowSignature", "_OriginalOrder", "_RowSeqWithinSignature", "_RowMatchKey",
]

BANNER_BLUE_HEADERS = set(DON_ONLY_BANNER_COLUMNS + [
    "Subject", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "Program {not in Banner}", "Catalog Title", "Section Title", "Crosslist Details {not in Banner}", "Campus {Ssasect Campus}", "Campus Restriction {Camp Restr CC}", "Schedule Type {not in Banner}", "Instruction Method {Inst Method}", "Level {Level SC}", "Credits {Sect Crs}", "Enrollment Cap {Max Enrl}", "Meeting Type", "Pre-work Start Date {not in Banner}", "Pre-work End Date {not in Banner}", "Course Start Date {Meet Start Date}", "Course End Date {Meet End Date}", "Post-work Start Date {not in Banner}", "Post-work End Date {not in Banner}", "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}", "SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT", "Room {Meet Room}", "Building {Meet Bldg}", "Instructor Name {Instr Name}", "Instructor Email {Instr Email}", "Instructor ID {Instr ID}", "% Responsibility", "load/contract {not in Banner}", "costs per credit {not in Banner}", "total costs {not in Banner}", "account to be charged {not in Banner}",
])

BANNER_GREY_HEADERS = {
    "CRN sorted {not in Banner}", "CRN", "X Lst", "Credit Catalog {Cat Crs}", "Semester Start Date {Soaterm Start Date}", "Semester End Date {Soaterm End Date}", "Fee Amount {Fees Amt}", "Fee Code {Detl Code}", "Fee Type", "Fee Level", "Fee Ind", "Fee Term", "SEM Department {Scacrse Dept}", "Activity Date", "Status", "Bill Hrs", "Sect Sch Type", "Prerequisites {Preq Areas}", "Link Ind", "Integration Code", "Link Conn", "Crs Attr", "Comments", "Major Restr", "Scacrse College", "Enrolled", "Waitlist Capacity", "Primary Ind", "Override", "1st Date Reg Opens", "Last Date Reg Opens", "OL Range From Date", "OL Range To Date", "OL Numb Units", "OL Duration Code", "Grade Mode", "Gradable Ind", "Tuit Waiver", "Course Override {Meet Override}", "Lvl Res Ind", "Cmp Res Ind", "Rate Code", "Cohort Code", "St Attr Code", "Coll SC", "DegC SC", "Prog SC", "FOS Type", "Special Approval {Special Apvl}", "Weeks", "Term Code", "Part-Of-Term", "Camp SC",
}


@dataclass(frozen=True)
class ComparisonResult:
    excel_bytes: bytes
    discrepancy_count: int
    compared_columns: int


def _unique_headers(headers) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for index, value in enumerate(headers, start=1):
        name = "" if value is None else str(value).strip()
        name = name or f"Unnamed_{index}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}_{count}")
    return result


def _normalize_key(series: pd.Series) -> pd.Series:
    value = series.astype(str).str.strip()
    return value.mask(value.str.lower().isin({"", "nan", "none", "null"}), np.nan)


def _normalize_value(value, column: str | None = None):
    if pd.isna(value) or (isinstance(value, str) and not value.strip()):
        return None
    if column in {"Building {Meet Bldg}", "Room {Meet Room}", "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}"}:
        if isinstance(value, str) and value.strip().lower().rstrip(".") == "arr":
            return None
    if column == "Schedule Type {not in Banner}" and isinstance(value, str) and value.strip().lower() in {"bl", "blended learning"}:
        return "BL"
    return value


def _equal(left, right, column: str) -> bool:
    left, right = _normalize_value(left, column), _normalize_value(right, column)
    if left is None or right is None:
        return left is right
    try:
        return float(left) == float(right)
    except (TypeError, ValueError):
        return str(left).strip() == str(right).strip()


def _load_donsheet(data: bytes) -> tuple[pd.DataFrame, list[str], object]:
    workbook = load_workbook(BytesIO(data), data_only=True)
    frames = []
    sheet_names = [name for name in workbook.sheetnames if name not in HELPER_SHEETS and not name.startswith("DropDown")]
    for name in sheet_names:
        raw = pd.DataFrame(workbook[name].values).dropna(axis=0, how="all").dropna(axis=1, how="all")
        if raw.empty:
            continue
        frame = raw.iloc[1:].copy()
        frame.columns = _unique_headers(raw.iloc[0].tolist())
        frame = frame.dropna(axis=0, how="all")
        if not frame.empty:
            frame["_SheetName"] = name
            frame["_ExcelRow"] = np.arange(2, len(frame) + 2)
            frames.append(frame)
    if not frames:
        raise ValueError("The DonSheet does not contain any usable course rows.")
    return pd.concat(frames, ignore_index=True), sheet_names, workbook


def _load_banner(data: bytes, filename: str) -> pd.DataFrame:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else "xlsx"
    if extension == "csv":
        frame = pd.read_csv(BytesIO(data), dtype=str)
    else:
        raw = pd.read_excel(BytesIO(data), header=None, nrows=25, dtype=str, na_filter=False)
        header_row = next(
            (index for index, row in raw.iterrows() if "CRN" in {str(value).strip() for value in row}),
            0,
        )
        frame = pd.read_excel(BytesIO(data), header=header_row, dtype=str, na_filter=False)
    frame.columns = _unique_headers(frame.columns)
    frame = frame.rename(columns=BANNER_RENAMES)
    if BASE_KEY not in frame.columns:
        raise ValueError("The Banner Sheet must contain a CRN column.")
    frame[BASE_KEY] = _normalize_key(frame[BASE_KEY])
    frame = frame[frame[BASE_KEY].notna()].copy()
    if "Building {Meet Bldg}" in frame.columns:
        frame["Building {Meet Bldg}"] = frame["Building {Meet Bldg}"].replace(SITE_NAMES)
    for column in DON_ONLY_BANNER_COLUMNS:
        if column not in frame.columns:
            frame[column] = np.nan
    relevant = [column for column in BANNER_RELEVANT_COLUMNS if column in frame.columns]
    if not relevant:
        relevant = [BASE_KEY]
    dedupe_view = frame[relevant].copy()
    for column in relevant:
        dedupe_view[column] = dedupe_view[column].map(lambda value: "" if _normalize_value(value) is None else str(value).strip())
    frame = frame.loc[~dedupe_view.duplicated(keep="first")].copy()
    sort_columns = [
        column for column in [
            "Meeting Type", "Course Start Date {Meet Start Date}", "Course End Date {Meet End Date}",
            "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}",
            "Instructor ID {Instr ID}", "% Responsibility", "Room {Meet Room}", "Building {Meet Bldg}",
            "SUN", "MON", "TUE", "WED", "THU", "FRI",
        ] if column in frame.columns
    ]
    frame = frame.sort_values([BASE_KEY, *sort_columns], kind="stable", na_position="last").copy()
    frame["CRN sorted {not in Banner}"] = (
        frame[BASE_KEY].astype(str).str.strip() + "-" + frame.groupby(BASE_KEY, dropna=False).cumcount().add(1).astype(str).str.zfill(2)
    )
    return frame.reset_index(drop=True)


def _add_match_key(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    usable = [column for column in HELPER_MATCH_COLUMNS if column in result.columns]
    if usable:
        result["_RowSignature"] = result[usable].fillna("").astype(str).apply(lambda row: " | ".join(value.strip() for value in row), axis=1)
    else:
        result["_RowSignature"] = "ROW"
    result["_OriginalOrder"] = np.arange(len(result))
    sort_columns = [BASE_KEY, "_RowSignature"]
    if "CRN sorted {not in Banner}" in result.columns:
        sort_columns.append("CRN sorted {not in Banner}")
    result = result.sort_values([*sort_columns, "_OriginalOrder"], kind="stable", na_position="last")
    result["_RowSequence"] = result.groupby([BASE_KEY, "_RowSignature"], dropna=False).cumcount().add(1)
    result["_RowMatchKey"] = result[BASE_KEY].astype(str) + " || " + result["_RowSignature"] + " || #" + result["_RowSequence"].astype(str)
    return result


def _record(frame: list[dict], row, kind: str, column: str, don_value=None, banner_value=None) -> None:
    frame.append({
        "SheetName": row.get("_SheetName", row.get("_SheetName_Don", row.get("Subject", "OnlyInBanner"))),
        "CRN": row.get("CRN", row.get("CRN_Don", row.get("CRN_Banner"))),
        "Subject": row.get("Subject", row.get("Subject_Don", row.get("Subject_Banner"))),
        "CRN sorted {not in Banner}": row.get("CRN sorted {not in Banner}", row.get("CRN sorted {not in Banner}_Don", row.get("CRN sorted {not in Banner}_Banner"))),
        "Course Number {Crse Num}": row.get("Course Number {Crse Num}", row.get("Course Number {Crse Num}_Don", row.get("Course Number {Crse Num}_Banner"))),
        "Course Section {Seq Crse Num}": row.get("Course Section {Seq Crse Num}", row.get("Course Section {Seq Crse Num}_Don", row.get("Course Section {Seq Crse Num}_Banner"))),
        "RowMatchKey": row.get("_RowMatchKey"), "Column": column,
        "DonSheet Value": don_value, "Banner Value": banner_value, "Type": kind,
    })


def generate_banner_donsheet_comparison(donsheet_bytes: bytes, banner_bytes: bytes, banner_filename: str) -> ComparisonResult:
    working, sheet_names, _ = _load_donsheet(donsheet_bytes)
    banner = _load_banner(banner_bytes, banner_filename)
    if BASE_KEY not in working.columns:
        raise ValueError("The DonSheet must contain a CRN column.")
    working[BASE_KEY] = _normalize_key(working[BASE_KEY])
    working = working[working[BASE_KEY].notna()].copy()
    # Banner exports include the entire university.  The notebook report limits
    # its comparison to the academic subjects represented in the DonSheet.
    if "Subject" in working.columns and "Subject" in banner.columns:
        subjects = set(working["Subject"].dropna().astype(str).str.strip())
        banner = banner[banner["Subject"].astype(str).str.strip().isin(subjects)].copy()
    working = _add_match_key(working)
    banner = _add_match_key(banner)
    pairs = [(don, banner_col) for don, banner_col in COMPARISON_COLUMNS.items() if don in working.columns and banner_col in banner.columns]
    if not pairs:
        raise ValueError("The uploaded files have no comparable course columns.")

    records: list[dict] = []
    working_crns, banner_crns = set(working[BASE_KEY]), set(banner[BASE_KEY])
    missing_in_banner, missing_in_donsheet = working_crns - banner_crns, banner_crns - working_crns
    shared_crns = working_crns & banner_crns
    working_shared = working[working[BASE_KEY].isin(shared_crns)].copy()
    banner_shared = banner[banner[BASE_KEY].isin(shared_crns)].copy()
    working_keys, banner_keys = set(working_shared["_RowMatchKey"]), set(banner_shared["_RowMatchKey"])

    for crn in sorted(missing_in_banner):
        row = working.loc[working[BASE_KEY] == crn].iloc[0]
        _record(records, row, "CRNMissingInBanner", "(entire CRN)", "(exists in Working only)", "(missing CRN)")
    for crn in sorted(missing_in_donsheet):
        row = banner.loc[banner[BASE_KEY] == crn].iloc[0]
        _record(records, row, "CRNMissingInDonSheet", "(entire CRN)", "(missing CRN)", "(exists in Banner only)")
    for key in sorted(working_keys - banner_keys):
        row = working_shared.loc[working_shared["_RowMatchKey"] == key].iloc[0]
        _record(records, row, "RowVariantMissingInBanner", "(row variant)", "(exists in Working only)", "(missing row variant)")
    for key in sorted(banner_keys - working_keys):
        row = banner_shared.loc[banner_shared["_RowMatchKey"] == key].iloc[0]
        _record(records, row, "RowVariantMissingInDonSheet", "(row variant)", "(missing row variant)", "(exists in Banner only)")

    working_merge_columns = [BASE_KEY, "_RowMatchKey", "_SheetName"] + (["CRN sorted {not in Banner}"] if "CRN sorted {not in Banner}" in working.columns else []) + [don for don, _ in pairs]
    banner_merge_columns = [BASE_KEY, "_RowMatchKey"] + [banner_column for _, banner_column in pairs]
    shared = working_shared[working_merge_columns].merge(
        banner_shared[banner_merge_columns], on=[BASE_KEY, "_RowMatchKey"], suffixes=("_Don", "_Banner"), how="inner"
    )
    for _, row in shared.iterrows():
        for don_col, banner_col in pairs:
            don_value = row.get(f"{don_col}_Don")
            banner_value = row.get(f"{banner_col}_Banner")
            if _equal(don_value, banner_value, don_col):
                continue
            kind = "Red" if _normalize_value(don_value, don_col) is not None and _normalize_value(banner_value, don_col) is None else "Green" if _normalize_value(don_value, don_col) is None else "Blue"
            _record(records, row, kind, don_col, don_value, banner_value)

    discrepancies = pd.DataFrame(records, columns=["SheetName", "Subject", "CRN sorted {not in Banner}", "CRN", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "RowMatchKey", "Column", "DonSheet Value", "Banner Value", "Type"])
    output = _build_report(donsheet_bytes, sheet_names, working, banner, discrepancies)
    return ComparisonResult(output, len(discrepancies), len(pairs))


def _build_report(source_bytes: bytes, sheet_names: list[str], working: pd.DataFrame, banner: pd.DataFrame, discrepancies: pd.DataFrame) -> bytes:
    # The notebook report copies calculated values, not live DonSheet formulas.
    # Keeping formulas can preserve external workbook references and make Excel
    # offer a repair/update prompt when the downloaded report is opened.
    source = load_workbook(BytesIO(source_bytes), data_only=True, keep_links=False)
    output = Workbook()
    output.remove(output.active)
    for sheet_name in sheet_names:
        if sheet_name not in source.sheetnames:
            continue
        original, target = source[sheet_name], output.create_sheet(sheet_name)
        for row in original.iter_rows():
            for cell in row:
                new_cell = target.cell(cell.row, cell.column, cell.value)
                if cell.has_style:
                    new_cell.font, new_cell.fill, new_cell.border = copy(cell.font), copy(cell.fill), copy(cell.border)
                    new_cell.alignment, new_cell.number_format, new_cell.protection = copy(cell.alignment), cell.number_format, copy(cell.protection)
        for column, dimension in original.column_dimensions.items():
            target.column_dimensions[column].width = dimension.width
        target.freeze_panes = original.freeze_panes

        header = {str(cell.value).strip(): cell.column for cell in target[1] if cell.value is not None}
        mismatching_column = header.get("Mismatching {not in Banner}")
        row_by_match_key = {
            row["_RowMatchKey"]: int(row["_ExcelRow"])
            for _, row in working.loc[working["_SheetName"] == sheet_name].iterrows()
        }
        crn_column = header.get(BASE_KEY)
        row_by_crn = {
            str(target.cell(row, crn_column).value).strip(): row
            for row in range(2, target.max_row + 1)
            if crn_column and target.cell(row, crn_column).value is not None
        }
        sheet_discrepancies = discrepancies[discrepancies["SheetName"] == sheet_name]
        for _, item in sheet_discrepancies.iterrows():
            target_row = row_by_match_key.get(item["RowMatchKey"])
            if target_row is None:
                target_row = row_by_crn.get(str(item["CRN"]).strip())
            if not target_row:
                continue
            if item["Type"] in {"CRNMissingInBanner", "RowVariantMissingInBanner"}:
                for column in range(1, target.max_column + 1):
                    target.cell(target_row, column).fill = DISCREPANCY_FILL
            elif item["Column"] in header:
                target.cell(target_row, header[item["Column"]]).fill = DISCREPANCY_FILL
                if mismatching_column is not None:
                    target.cell(target_row, mismatching_column).fill = DISCREPANCY_FILL
        target.freeze_panes = "H2"
        target.auto_filter.ref = f"A1:{target.cell(1, target.max_column).coordinate[:-1]}{target.max_row}"

    only_in_banner = discrepancies[discrepancies["Type"].isin({"CRNMissingInDonSheet", "RowVariantMissingInDonSheet"})]
    if not only_in_banner.empty:
        worksheet = output.create_sheet("OnlyInBanner")
        worksheet.sheet_properties.tabColor = "FFFF0000"
        crns = set(only_in_banner["CRN"].astype(str))
        rows = banner[banner[BASE_KEY].astype(str).isin(crns)].copy()
        rows = rows.rename(columns={"_RowSequence": "_RowSeqWithinSignature"})
        for column in BANNER_OUTPUT_COLUMNS:
            if column not in rows.columns:
                rows[column] = np.nan
        rows = rows[BANNER_OUTPUT_COLUMNS]
        for row in dataframe_to_rows(rows, index=False, header=True):
            worksheet.append(row)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = DISCREPANCY_FILL
        worksheet.row_dimensions[1].height = 150
        for cell in worksheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True, color="FF000000")
            cell.fill = (
                SUMMARY_HEADER_BLUE_FILL if cell.value in BANNER_BLUE_HEADERS
                else SUMMARY_HEADER_GREY_FILL if cell.value in BANNER_GREY_HEADERS
                else SUMMARY_HEADER_DEFAULT_FILL
            )
            cell.alignment = Alignment(horizontal="center", textRotation=90, wrap_text=True)
        fixed_widths = {
            "Notes {not in Banner}": 40, "Instructor Name {Instr Name}": 20,
            "Instructor Email {Instr Email}": 20, "Instructor ID {Instr ID}": 10,
            "account to be charged {not in Banner}": 20,
        }
        for name, width in fixed_widths.items():
            if name in BANNER_OUTPUT_COLUMNS:
                worksheet.column_dimensions[get_column_letter(BANNER_OUTPUT_COLUMNS.index(name) + 1)].width = width
        worksheet.freeze_panes = "H3"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, worksheet.max_column).coordinate[:-1]}{worksheet.max_row}"

    summary = output.create_sheet("SummaryReport")
    summary.append(["Sheet", "Subject", "CRN sorted {not in Banner}", "CRN", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "Column", "Type", "Banner Value", "DonSheet Value", "Status"])
    summary_data = discrepancies.copy()
    summary_data["Status"] = summary_data["Type"].map({
        "Red": "Value missing in Banner", "Green": "Value missing in DonSheet", "Blue": "Values differ",
        "CRNMissingInBanner": "CRN missing in Banner", "CRNMissingInDonSheet": "CRN missing in DonSheet",
        "RowVariantMissingInBanner": "Row variant missing in Banner", "RowVariantMissingInDonSheet": "Row variant missing in DonSheet",
    }).fillna("")
    summary_data = summary_data.rename(columns={"SheetName": "Sheet"})
    for row in dataframe_to_rows(summary_data[["Sheet", "Subject", "CRN sorted {not in Banner}", "CRN", "Course Number {Crse Num}", "Course Section {Seq Crse Num}", "Column", "Type", "Banner Value", "DonSheet Value", "Status"]], index=False, header=False):
        summary.append(row)
    header_fills = {
        "Subject": SUMMARY_HEADER_BLUE_FILL,
        "Course Number {Crse Num}": SUMMARY_HEADER_BLUE_FILL,
        "Course Section {Seq Crse Num}": SUMMARY_HEADER_BLUE_FILL,
        "CRN sorted {not in Banner}": SUMMARY_HEADER_GREY_FILL,
        "CRN": SUMMARY_HEADER_GREY_FILL,
        "Status": SUMMARY_HEADER_GREY_FILL,
    }
    summary.row_dimensions[1].height = 150
    for cell in summary[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color="FF000000")
        cell.fill = header_fills.get(cell.value, SUMMARY_HEADER_DEFAULT_FILL)
        cell.alignment = Alignment(horizontal="center", textRotation=90, wrap_text=True)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:K{summary.max_row}"
    stream = BytesIO()
    output.save(stream)
    return stream.getvalue()
