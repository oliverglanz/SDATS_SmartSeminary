from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook as load_openpyxl_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.worksheet.worksheet import Worksheet

from .config import HELPER_SHEETS
from .models import WorkbookSummary


def _read_source_bytes(source: str | Path | bytes | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, Path):
        return source.read_bytes()
    if isinstance(source, str):
        return Path(source).read_bytes()
    if hasattr(source, "read"):
        content = source.read()
        if hasattr(source, "seek"):
            source.seek(0)
        return content
    return Path(source).read_bytes()


def _load_excel_file(source: str | Path | bytes | BinaryIO) -> pd.ExcelFile:
    return pd.ExcelFile(BytesIO(_read_source_bytes(source)))


def _load_openpyxl_workbook(source: str | Path | bytes | BinaryIO):
    return load_openpyxl_workbook(BytesIO(_read_source_bytes(source)), data_only=True)


def _color_to_rgb(color) -> str | None:
    if color is None or color.type is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        return rgb[-6:].upper()
    if color.type == "indexed" and color.indexed is not None:
        indexed = COLOR_INDEX[color.indexed]
        return str(indexed)[-6:].upper()
    return None


def _is_red_rgb(rgb: str | None) -> bool:
    if not rgb or len(rgb) != 6:
        return False
    try:
        red = int(rgb[0:2], 16)
        green = int(rgb[2:4], 16)
        blue = int(rgb[4:6], 16)
    except ValueError:
        return False
    return red >= 180 and green <= 120 and blue <= 120


def _cell_has_red_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    return _is_red_rgb(_color_to_rgb(fill.fgColor)) or _is_red_rgb(_color_to_rgb(fill.start_color))


def _red_highlighted_row_indices(sheet: Worksheet, column_count: int) -> set[int]:
    red_rows: set[int] = set()
    for row_number in range(2, sheet.max_row + 1):
        row_cells = sheet[row_number][: max(column_count, 1)]
        if any(_cell_has_red_fill(cell) for cell in row_cells):
            red_rows.add(row_number - 2)
    return red_rows


def _source_name(source: str | Path | bytes | BinaryIO, excel_file: pd.ExcelFile) -> str:
    explicit_name = getattr(source, "name", None)
    if explicit_name:
        return str(explicit_name)
    if isinstance(source, Path):
        return source.name
    if isinstance(source, str):
        return Path(source).name
    excel_io = getattr(excel_file, "io", None)
    if isinstance(excel_io, (str, Path)):
        return Path(excel_io).name
    return "uploaded workbook"


def load_workbook(source: str | Path | bytes | BinaryIO) -> tuple[pd.ExcelFile, WorkbookSummary]:
    excel_file = _load_excel_file(source)
    sheet_names = list(excel_file.sheet_names)
    data_sheet_names = [name for name in sheet_names if name not in HELPER_SHEETS]
    helper_sheet_names = [name for name in sheet_names if name in HELPER_SHEETS]

    row_counts: dict[str, int] = {}
    warnings: list[str] = []

    for sheet_name in data_sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        row_counts[sheet_name] = len(df)
        if df.empty:
            warnings.append(f"Sheet '{sheet_name}' is empty.")

    summary = WorkbookSummary(
        source_name=_source_name(source, excel_file),
        sheet_names=sheet_names,
        data_sheet_names=data_sheet_names,
        helper_sheet_names=helper_sheet_names,
        row_counts=row_counts,
        warnings=warnings,
    )
    return excel_file, summary


def read_department_sheets(
    excel_file: pd.ExcelFile,
    sheet_names: list[str],
    source: str | Path | bytes | BinaryIO | None = None,
) -> dict[str, pd.DataFrame]:
    department_frames: dict[str, pd.DataFrame] = {}
    styled_workbook = _load_openpyxl_workbook(source) if source is not None else None
    for sheet_name in sheet_names:
        frame = pd.read_excel(excel_file, sheet_name=sheet_name)
        if styled_workbook is not None and sheet_name in styled_workbook.sheetnames and not frame.empty:
            red_rows = _red_highlighted_row_indices(styled_workbook[sheet_name], len(frame.columns))
            if red_rows:
                frame = frame.drop(index=[row for row in red_rows if row in frame.index]).reset_index(drop=True)
        department_frames[sheet_name] = frame
    return department_frames


def read_helper_sheet(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in excel_file.sheet_names:
        return pd.DataFrame()
    return pd.read_excel(excel_file, sheet_name=sheet_name)
