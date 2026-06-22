from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A2, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


IGNORE_SHEETS = {"DropDownMenu", "InstructorMap"}
BLOCK_ORDER = [
    "Monday Morning",
    "Tuesday Morning",
    "Tuesday 11:30 - 12:30 Chapel",
    "Wednesday Morning",
    "Thursday Morning",
    "Lunch 12:20-2:20 PM / PATH588 Seminary Chorus (CRN 498): 1:30-2:20 PM",
    "Monday Afternoon",
    "Tuesday Afternoon",
    "Wednesday Afternoon",
    "Thursday Afternoon",
    "Supper 5:30 - 6:20 PM",
    "Evening Classes",
]
SPECIAL_ROWS = {
    "Tuesday 11:30 - 12:30 Chapel": "Chapel reserved time",
    "Lunch 12:20-2:20 PM / PATH588 Seminary Chorus (CRN 498): 1:30-2:20 PM": "Lunch window / PATH588 Seminary Chorus if scheduled",
    "Supper 5:30 - 6:20 PM": "Supper reserved time",
}
TERM_SEQUENCE = ("Term 1", "Term 2", "Term 3")
SUBJECT_FILLS = {
    "OTST": "#fce5cd",
    "ANEA": "#fce5cd",
    "NTST": "#ff9900",
    "PATH": "#00ffff",
    "THST": "#ea9999",
    "DSLE": "#00ff00",
    "GSEM": "#ff8aff",
    "CHIS": "#d9d2e9",
    "MSSN": "#a4c2f4",
}
ROOM_ORDER = ["N108", "N110", "N135", "N150", "S115", "S120", "N215", "N235", "N310", "N335", "S340"]
ROOM_HEADER_FILL = "#ffff00"
TITLE_FILL_SYNC = "#b6d7a8"
TITLE_FILL_SYNC_ASYNC = "#ffe599"
TIMEBLOCK_FILL = "#ffff00"
SPECIAL_ROW_FILL = "#d9d9d9"
TERM_HEADER_FILL = "#666666"
COURSE_LIST_PATH = Path(__file__).resolve().parents[2] / "0_source_files" / "0000_BuildingFiles" / "source_ClassAttributes_v20260511.xlsx"
LOCAL_TIMEZONE = ZoneInfo("America/Detroit")


@dataclass
class ScheduleGenerationResult:
    preview_df: pd.DataFrame
    normalized_meetings_df: pd.DataFrame
    excluded_df: pd.DataFrame
    excel_bytes: bytes
    pdf_bytes: bytes
    excel_filename: str
    pdf_filename: str


def generate_schedule_outputs(source_name: str, uploaded_bytes: bytes) -> ScheduleGenerationResult:
    df = _load_raw_donsheet(uploaded_bytes)
    title_method_map = _build_title_method_map(df)
    course_annotation_map = _load_course_annotation_map()
    async_report_df = _build_async_report_df(df, course_annotation_map)
    meetings, excluded_df = _build_meetings(df)
    slots = _group_slots(meetings)
    summer_mode = _summer_term_mode(df)
    block_order, special_rows = _summer_block_order(meetings) if summer_mode else (BLOCK_ORDER, SPECIAL_ROWS)
    term_summary_map = _build_term_summary_map(meetings, block_order)
    sync_meetings = meetings[
        meetings.apply(lambda row: _slot_has_sync(row.get("Course Code"), row.get("Title"), title_method_map), axis=1)
    ].copy() if not meetings.empty else meetings.copy()
    sync_slots = _group_slots(sync_meetings)
    sync_block_order, sync_special_rows = _summer_block_order(sync_meetings) if summer_mode else (BLOCK_ORDER, SPECIAL_ROWS)
    sync_term_summary_map = _build_term_summary_map(sync_meetings, sync_block_order)
    excel_bytes = _build_schedule_excel(
        slots,
        sync_slots,
        meetings,
        excluded_df,
        async_report_df,
        title_method_map,
        course_annotation_map,
        block_order,
        special_rows,
        term_summary_map,
        sync_block_order,
        sync_special_rows,
        sync_term_summary_map,
    )
    pdf_bytes = _build_schedule_pdf(
        slots,
        sync_slots,
        async_report_df,
        title_method_map,
        course_annotation_map,
        block_order,
        special_rows,
        term_summary_map,
        sync_block_order,
        sync_special_rows,
        sync_term_summary_map,
    )
    stem = _safe_stem(source_name)
    return ScheduleGenerationResult(
        preview_df=slots,
        normalized_meetings_df=meetings,
        excluded_df=excluded_df,
        excel_bytes=excel_bytes,
        pdf_bytes=pdf_bytes,
        excel_filename=f"{stem}_weekly_schedule.xlsx",
        pdf_filename=f"{stem}_weekly_schedule.pdf",
    )


def _load_raw_donsheet(uploaded_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(BytesIO(uploaded_bytes), data_only=True)
    records: list[dict[str, object]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if row.get("CRN") is None and row.get("Catalog Title") is None and row.get("Section Title") is None:
                continue
            row["_sheet"] = sheet_name
            records.append(row)
    return pd.DataFrame(records)


def _normalize_inst_method(value: object) -> str:
    return str(value or "").strip().upper()


def _cell_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def _normalize_time_token(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%H%M")
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if len(digits) in {3, 4}:
        return digits.zfill(4)
    return ""


def _parse_excelish_date(value: object) -> pd.Timestamp | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed


def _is_valid_time(value: object) -> bool:
    return bool(re.fullmatch(r"\d{4}", _normalize_time_token(value)))


def _fmt_time_ampm(value: object) -> str:
    s = _normalize_time_token(value)
    hh, mm = int(s[:2]), int(s[2:])
    ampm = "AM" if hh < 12 else "PM"
    hh12 = hh % 12 or 12
    return f"{hh12}:{mm:02d} {ampm}"


def _fmt_range_ampm(start: object, end: object) -> str:
    return f"{_fmt_time_ampm(start)}-{_fmt_time_ampm(end)}"


def _fmt_time_military(value: object) -> str:
    s = _normalize_time_token(value)
    return f"{s[:2]}:{s[2:]}"


def _fmt_range_short(start: object, end: object) -> str:
    return f"{_fmt_time_military(start)}-{_fmt_time_military(end)}"


def _uniq(seq) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for item in seq:
        if pd.isna(item):
            continue
        text = str(item).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def _pick_title(row: pd.Series) -> str:
    sec = _cell_text(row.get("Section Title"))
    cat = _cell_text(row.get("Catalog Title"))
    if sec and sec.lower() not in {"none", cat.lower()} and sec != "0":
        return sec
    return cat


def _normalize_room(room: object) -> str:
    return re.sub(r"\s+", "", str(room or "").strip()).upper()


def _display_room(room: object) -> str:
    normalized = _normalize_room(room)
    match = re.match(r"([A-Z]+)(\d+)", normalized)
    return f"{match.group(1)} {match.group(2)}" if match else str(room or "").strip()


def _room_sort_key(room: str):
    normalized = _normalize_room(room)
    if normalized in ROOM_ORDER:
        return (0, ROOM_ORDER.index(normalized))
    match = re.match(r"([A-Z]+)(\d+)", normalized)
    return (1, match.group(1), int(match.group(2))) if match else (2, normalized, 0)


def _slot_subject(codes: str) -> str:
    match = re.match(r"([A-Z]+)", str(codes or ""))
    return match.group(1) if match else ""


def _build_course_key(subject: object, number: object, section: object) -> str:
    subject_text = _cell_text(subject)
    number_text = _cell_text(number)
    section_text = _cell_text(section)
    if not subject_text and not number_text:
        return ""
    key = f"{subject_text} {number_text}".strip()
    return f"{key}-{section_text}" if section_text else key


def _build_course_family_key(subject: object, number: object) -> str:
    subject_text = _cell_text(subject)
    number_text = _cell_text(number)
    if not subject_text and not number_text:
        return ""
    return f"{subject_text} {number_text}".strip()


def _first_course_key(course_codes: object) -> str:
    return str(course_codes or "").split(" / ")[0].strip()


def _course_subject_from_key(course_key: str) -> str:
    match = re.match(r"([A-Z]+)", str(course_key or ""))
    return match.group(1) if match else ""


def _course_family_from_key(course_key: object) -> str:
    match = re.match(r"^([A-Z]+)\s+([0-9A-Z]+)", str(course_key or "").strip())
    return f"{match.group(1)} {match.group(2)}" if match else ""


def _normalize_course_family_text(text: object) -> str:
    match = re.match(r"^([A-Z]+)\s*([0-9A-Z]+)\s*$", _cell_text(text).upper())
    return f"{match.group(1)} {match.group(2)}" if match else ""


def _load_course_annotation_map() -> dict[str, str]:
    if not COURSE_LIST_PATH.exists():
        return {}
    try:
        wb = load_workbook(COURSE_LIST_PATH, data_only=True)
    except Exception:
        return {}
    if "Course List" not in wb.sheetnames:
        return {}
    ws = wb["Course List"]
    annotations: dict[str, str] = {}
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        first = _cell_text(row[0] if row else "")
        attribute = _cell_text(row[2] if row and len(row) > 2 else "")
        if not first or not attribute:
            continue
        course_key = _normalize_course_family_text(first)
        if not course_key:
            continue
        annotations[course_key] = attribute
    return annotations


def _course_annotation_for_slot(course_codes: object, annotation_map: dict[str, str]) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for course_key in [part.strip() for part in str(course_codes or "").split(" / ") if part.strip()]:
        family_key = _course_family_from_key(course_key)
        value = annotation_map.get(family_key, "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return " / ".join(values)


def _is_prerequisite_annotation(text: object) -> bool:
    return "prerequisite" in str(text or "").strip().lower()


def _printed_on_text() -> str:
    return f"Printed on {datetime.now(LOCAL_TIMEZONE).strftime('%B %d, %Y')}"


def _build_async_report_df(df: pd.DataFrame, course_annotation_map: dict[str, str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Course / Sect.", "CRN #", "Course Title", "Attribute/Notes", "Cr.", "Instructor", "Cap"])
    async_df = df.copy()
    async_df["Inst Method Norm"] = async_df["Instruction Method {Inst Method}"].map(_normalize_inst_method)
    async_df = async_df[async_df["Inst Method Norm"] == "ASYNC"].copy()
    if async_df.empty:
        return pd.DataFrame(columns=["Course / Sect.", "CRN #", "Course Title", "Attribute/Notes", "Cr.", "Instructor", "Cap"])
    async_df["Course Key"] = async_df.apply(
        lambda row: _build_course_key(
            row.get("Subject"),
            row.get("Course Number {Crse Num}"),
            row.get("Course Section {Seq Crse Num}"),
        ),
        axis=1,
    )
    async_df["Course Family"] = async_df.apply(
        lambda row: _build_course_family_key(row.get("Subject"), row.get("Course Number {Crse Num}")),
        axis=1,
    )
    async_df["Subject Code"] = async_df["Subject"].map(_cell_text)
    async_df["Course / Sect."] = async_df["Course Key"].map(lambda v: str(v).replace(" ", ""))
    async_df["CRN #"] = async_df["CRN"].map(_cell_text)
    async_df["Course Title"] = async_df.apply(_pick_title, axis=1)
    async_df["Attribute/Notes"] = async_df["Course Family"].map(lambda key: course_annotation_map.get(key, ""))
    async_df["Cr."] = async_df["Credits {Sect Crs}"].map(lambda v: _normalize_enrollment(v))
    async_df["Instructor"] = async_df["Instructor Name {Instr Name}"].map(_cell_text)
    async_df["Cap"] = async_df["Enrollment Cap {Max Enrl}"].map(_normalize_enrollment)
    async_df["_course_num_sort"] = async_df["Course Number {Crse Num}"].map(lambda v: int(re.sub(r"\D", "", _cell_text(v)) or 0))
    async_df["_section_sort"] = async_df["Course Section {Seq Crse Num}"].map(_cell_text)
    async_df = async_df.sort_values(["Subject", "_course_num_sort", "_section_sort", "CRN #", "Course Title"]).drop_duplicates(
        subset=["Course / Sect.", "CRN #", "Course Title", "Attribute/Notes", "Cr.", "Instructor", "Cap"]
    )
    return async_df[["Course / Sect.", "CRN #", "Course Title", "Attribute/Notes", "Cr.", "Instructor", "Cap", "Course Key", "Subject Code"]].reset_index(drop=True)


def _normalize_enrollment(value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _title_fill_for_methods(methods: set[str]) -> str | None:
    methods = {method for method in methods if method}
    if {"INPERSON", "SYNC", "ASYNC"}.issubset(methods):
        return TITLE_FILL_SYNC_ASYNC
    if "INPERSON" in methods and "SYNC" in methods:
        return TITLE_FILL_SYNC
    return None


def _method_value_for_aggregation(row: pd.Series) -> str:
    method = _normalize_inst_method(row.get("Instruction Method {Inst Method}"))
    if method:
        return method
    has_room = bool(_cell_text(row.get("Room {Meet Room}")))
    has_start = _is_valid_time(row.get("Course Beginning Time {Meet Beg Time}"))
    has_end = _is_valid_time(row.get("Course Ending Time {Meet End Time}"))
    has_meeting_pattern = any(_cell_text(row.get(col)) for col in ("MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"))
    if has_room and has_start and has_end and has_meeting_pattern:
        return "INPERSON"
    return ""


def _build_day_pattern(row: pd.Series) -> str:
    day_cols = [("MON", "M"), ("TUE", "T"), ("WED", "W"), ("THU", "R"), ("FRI", "F"), ("SAT", "S"), ("SUN", "U")]
    marks: list[str] = []
    for col, mark in day_cols:
        if str(row.get(col) or "").strip() == mark:
            marks.append(mark)
    return "".join(marks)


def _term_from_month(month: int | None) -> str:
    if month == 4:
        return "Term 1"
    if month == 5:
        return "Term 1"
    if month == 6:
        return "Term 2"
    if month == 7:
        return "Term 3"
    return ""


def _summer_term_mode(df: pd.DataFrame) -> bool:
    if df.empty or "Semester Start Date {Soaterm Start Date}" not in df.columns:
        return False
    months = {
        parsed.month
        for parsed in (_parse_excelish_date(value) for value in df["Semester Start Date {Soaterm Start Date}"])
        if parsed is not None
    }
    return bool(months & {4, 5})


def _resolve_course_term(row: pd.Series) -> str:
    for field in (
        "Course Start Date {Meet Start Date}",
        "Meet Start Date",
        "Semester Start Date {Soaterm Start Date}",
    ):
        parsed = _parse_excelish_date(row.get(field))
        if parsed is not None:
            term = _term_from_month(parsed.month)
            if term:
                return term
    return ""


def _summer_block_order(meetings: pd.DataFrame) -> tuple[list[str], dict[str, str]]:
    present_terms = [term for term in TERM_SEQUENCE if not meetings.empty and term in set(meetings.get("Term", []))]
    if not present_terms:
        present_terms = list(TERM_SEQUENCE)
    order: list[str] = []
    specials: dict[str, str] = {}
    for term in present_terms:
        header_key = f"{term} Header"
        chapel_key = f"{term} Chapel"
        lunch_key = f"{term} Lunch"
        order.extend([header_key, f"{term} Morning", chapel_key, lunch_key, f"{term} Afternoon"])
        specials[header_key] = ""
        specials[chapel_key] = "Chapel reserved time"
        specials[lunch_key] = "Lunch window / PATH588 Seminary Chorus if scheduled"
    return order, specials


def _format_term_summary_label(term: str, meetings: pd.DataFrame) -> str:
    if meetings.empty or "Term" not in meetings.columns:
        return term
    term_rows = meetings[meetings["Term"].astype(str).str.strip() == term].copy()
    if term_rows.empty:
        return term
    start_dates = [
        parsed for parsed in (_parse_excelish_date(value) for value in term_rows.get("Course Start Date", []))
        if parsed is not None
    ]
    end_dates = [
        parsed for parsed in (_parse_excelish_date(value) for value in term_rows.get("Course End Date", []))
        if parsed is not None
    ]
    if not start_dates or not end_dates:
        return term
    start_date = min(start_dates)
    end_date = max(end_dates)
    weeks = max(1, int(((end_date.normalize() - start_date.normalize()).days + 1 + 6) // 7))
    return f"{term} ({weeks} weeks) {start_date.strftime('%b')} {start_date.day} to {end_date.strftime('%b')} {end_date.day}"


def _build_term_summary_map(meetings: pd.DataFrame, block_order: list[str]) -> dict[str, str]:
    summaries: dict[str, str] = {}
    for block in block_order:
        if block.endswith(" Header"):
            term = block.removesuffix(" Header")
            summaries[block] = _format_term_summary_label(term, meetings)
    return summaries


def _method_set_for_slot(course_codes: object, title_text: object, method_map: dict[str, set[str]]) -> set[str]:
    course_keys = [part.strip() for part in str(course_codes or "").split(" / ") if part.strip()]
    combined: set[str] = set()
    for course_key in course_keys:
        combined |= method_map.get(course_key, set())
        family_key = _course_family_from_key(course_key)
        if family_key:
            combined |= method_map.get(family_key, set())
    titles = [part.strip() for part in str(title_text or "").split(" / ") if part.strip()]
    for title in titles:
        combined |= method_map.get(title, set())
    return combined


def _slot_has_sync(course_codes: object, title_text: object, method_map: dict[str, set[str]]) -> bool:
    return "SYNC" in _method_set_for_slot(course_codes, title_text, method_map)


def _assign_block(day: str, start: object, end: object, *, term: str = "", summer_mode: bool = False) -> str | None:
    s = int(_normalize_time_token(start))
    e = int(_normalize_time_token(end))
    if summer_mode and term:
        if s < 1130:
            return f"{term} Morning"
        if s < 1730:
            return f"{term} Afternoon"
        return f"{term} Evening"
    if day == "Monday":
        return "Monday Morning" if s < 1220 else "Monday Afternoon" if s < 1800 else "Evening Classes"
    if day == "Tuesday":
        if s < 1130:
            return "Tuesday Morning"
        if 1130 <= s < 1230 or (s < 1230 and e > 1130):
            return "Tuesday 11:30 - 12:30 Chapel"
        return "Tuesday Afternoon" if s < 1800 else "Evening Classes"
    if day == "Wednesday":
        return "Wednesday Morning" if s < 1220 else "Wednesday Afternoon" if s < 1800 else "Evening Classes"
    if day == "Thursday":
        return "Thursday Morning" if s < 1220 else "Thursday Afternoon" if s < 1800 else "Evening Classes"
    return None


def _build_title_method_map(df: pd.DataFrame) -> dict[str, set[str]]:
    if df.empty:
        return {}
    title_df = df.copy()
    title_df["Printable Title"] = title_df.apply(_pick_title, axis=1)
    title_df["Course Key"] = title_df.apply(
        lambda row: _build_course_key(
            row.get("Subject"),
            row.get("Course Number {Crse Num}"),
            row.get("Course Section {Seq Crse Num}"),
        ),
        axis=1,
    )
    title_df["Course Family"] = title_df.apply(
        lambda row: _build_course_family_key(row.get("Subject"), row.get("Course Number {Crse Num}")),
        axis=1,
    )
    title_df["Catalog Title Key"] = title_df["Catalog Title"].map(_cell_text)
    title_df["Section Title Key"] = title_df["Section Title"].map(_cell_text)
    title_df["Instruction Method {Inst Method}"] = title_df.apply(_method_value_for_aggregation, axis=1)
    method_map: dict[str, set[str]] = {}
    for key_name in ("Printable Title", "Catalog Title Key", "Section Title Key", "Course Key", "Course Family"):
        for key, group in title_df.groupby(key_name, dropna=False):
            key_text = _cell_text(key)
            if not key_text:
                continue
            methods = {m for m in group["Instruction Method {Inst Method}"].tolist() if m}
            if key_text in method_map:
                method_map[key_text] |= methods
            else:
                method_map[key_text] = methods
    return method_map


def _build_meetings(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        empty = pd.DataFrame()
        return empty, empty
    summer_mode = _summer_term_mode(df)
    dedupe_cols = [
        "CRN", "Subject", "Course Number {Crse Num}", "Course Section {Seq Crse Num}",
        "Catalog Title", "Section Title", "Credits {Sect Crs}",
        "Course Beginning Time {Meet Beg Time}", "Course Ending Time {Meet End Time}",
        "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN",
        "Room {Meet Room}", "Instructor Name {Instr Name}", "Instruction Method {Inst Method}",
    ]
    valid = df[
        df["Course Beginning Time {Meet Beg Time}"].map(_is_valid_time)
        & df["Course Ending Time {Meet End Time}"].map(_is_valid_time)
        & df["Room {Meet Room}"].notna()
        & df["Subject"].notna()
    ].copy().drop_duplicates(subset=[c for c in dedupe_cols if c in df.columns])
    valid["Normalized Room"] = valid["Room {Meet Room}"].map(_normalize_room)
    valid = valid[valid["Normalized Room"].isin(ROOM_ORDER)].copy()

    day_map = [("MON", "Monday", "M"), ("TUE", "Tuesday", "T"), ("WED", "Wednesday", "W"), ("THU", "Thursday", "R")]
    meeting_rows: list[dict[str, object]] = []
    for _, row in valid.iterrows():
        day_pattern = _build_day_pattern(row)
        for day_col, day_name, marker in day_map:
            if str(row.get(day_col) or "").strip() == marker:
                section = _cell_text(row.get("Course Section {Seq Crse Num}"))
                subject = _cell_text(row.get("Subject"))
                number = _cell_text(row.get("Course Number {Crse Num}"))
                course_key = _build_course_key(subject, number, section)
                term = _resolve_course_term(row) if summer_mode else ""
                meeting_rows.append(
                    {
                        "Day": day_name,
                        "Room": _normalize_room(row["Room {Meet Room}"]),
                        "Room Display": _display_room(row["Room {Meet Room}"]),
                        "Start": _normalize_time_token(row["Course Beginning Time {Meet Beg Time}"]),
                        "End": _normalize_time_token(row["Course Ending Time {Meet End Time}"]),
                        "CRN": _cell_text(row.get("CRN")),
                        "Subject": subject,
                        "Course Number": number,
                        "Section": section,
                        "Course Code": course_key,
                        "Title": _pick_title(row),
                        "Catalog Title": _cell_text(row.get("Catalog Title")),
                        "Instructor": _cell_text(row.get("Instructor Name {Instr Name}")),
                        "Credits": _cell_text(row.get("Credits {Sect Crs}")),
                        "Enrollment Cap": _normalize_enrollment(row.get("Enrollment Cap {Max Enrl}")),
                        "Term": term,
                        "Course Start Date": row.get("Course Start Date {Meet Start Date}", row.get("Meet Start Date")),
                        "Course End Date": row.get("Course End Date {Meet End Date}", row.get("Meet End Date")),
                        "Day Pattern": day_pattern,
                        "DayTime": f"{day_pattern} {_fmt_range_short(row['Course Beginning Time {Meet Beg Time}'], row['Course Ending Time {Meet End Time}'])}".strip(),
                        "Inst Method": _normalize_inst_method(row.get("Instruction Method {Inst Method}")),
                    }
                )
    meetings = pd.DataFrame(meeting_rows)
    excluded_rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        reasons: list[str] = []
        normalized_room = _normalize_room(row.get("Room {Meet Room}"))
        if pd.isna(row.get("Room {Meet Room}")) or str(row.get("Room {Meet Room}") or "").strip() == "":
            reasons.append("No room")
        elif normalized_room not in ROOM_ORDER:
            reasons.append("Room not shown in printable schedule")
        if not _is_valid_time(row.get("Course Beginning Time {Meet Beg Time}")) or not _is_valid_time(row.get("Course Ending Time {Meet End Time}")):
            reasons.append("Invalid or missing time")
        if reasons:
            excluded_rows.append(
                {
                    "CRN": row.get("CRN"),
                    "Subject": row.get("Subject"),
                    "Course Number {Crse Num}": row.get("Course Number {Crse Num}"),
                    "Course Section {Seq Crse Num}": row.get("Course Section {Seq Crse Num}"),
                    "Catalog Title": row.get("Catalog Title"),
                "Room {Meet Room}": _display_room(row.get("Room {Meet Room}")),
                    "Course Beginning Time {Meet Beg Time}": row.get("Course Beginning Time {Meet Beg Time}"),
                    "Course Ending Time {Meet End Time}": row.get("Course Ending Time {Meet End Time}"),
                    "Reason": "; ".join(reasons),
                }
            )
    excluded_df = pd.DataFrame(excluded_rows)
    return meetings, excluded_df


def _group_slots(meetings: pd.DataFrame) -> pd.DataFrame:
    if meetings.empty:
        return pd.DataFrame(columns=["Day", "Room", "Room Display", "Start", "End", "Course Codes", "Title", "Instructor", "CRNs", "Credits", "Enrollment Cap", "Term", "DayTime", "Time", "Block"])
    summer_mode = "Term" in meetings.columns and meetings["Term"].astype(str).str.strip().ne("").any()
    grouped_rows: list[dict[str, object]] = []
    group_fields = ["Term", "Room", "Start", "End"] if summer_mode else ["Day", "Room", "Start", "End"]
    for keys, group in meetings.groupby(group_fields, sort=False):
        group = group.reset_index(drop=True)
        primary_row = group.iloc[0]
        codes = _uniq(group["Course Code"])
        titles = _uniq(group["Title"])
        catalog_titles = _uniq(group["Catalog Title"])
        if len(titles) > 2:
            titles = catalog_titles[:2]
        if summer_mode:
            day_value = ""
            term_value, room_value, start_value, end_value = keys
        else:
            day_value, room_value, start_value, end_value = keys
            term_value = str(primary_row.get("Term") or "").strip()
        grouped_rows.append(
            {
                "Day": day_value,
                "Room": room_value,
                "Room Display": primary_row["Room Display"],
                "Start": start_value,
                "End": end_value,
                "Course Codes": " / ".join(codes),
                "Title": " / ".join(titles),
                "Instructor": " / ".join(_uniq(group["Instructor"])),
                "CRNs": ", ".join([c for c in _uniq(group["CRN"]) if c not in {"###", "####"}]),
                "Credits": " / ".join(_uniq(group["Credits"])),
                "Enrollment Cap": str(primary_row.get("Enrollment Cap") or "").strip(),
                "Credit CRN Lines": _build_credit_crn_lines(group),
                "Term": term_value,
                "DayTime": " / ".join(_uniq(group["DayTime"])),
                "Time": _fmt_range_ampm(start_value, end_value),
            }
        )
    slots = pd.DataFrame(grouped_rows)
    if summer_mode and not slots.empty:
        term_order = {term: index for index, term in enumerate(TERM_SEQUENCE)}
        slots["_term_order"] = slots["Term"].map(lambda value: term_order.get(value, 999))
        slots = slots.sort_values(["_term_order", "Start", "Room", "Course Codes"]).drop(columns="_term_order").reset_index(drop=True)
    else:
        slots = slots.sort_values(["Day", "Start", "Room", "Course Codes"]).reset_index(drop=True)
    slots["Block"] = slots.apply(
        lambda row: _assign_block(
            row["Day"],
            row["Start"],
            row["End"],
            term=str(row.get("Term") or "").strip(),
            summer_mode=summer_mode,
        ),
        axis=1,
    )
    return slots


def _sync_display_blocks(block_order: list[str], special_rows: dict[str, str]) -> list[str]:
    return [block for block in block_order if block.endswith(" Header") or block not in special_rows]


def _sync_block_time_groups(sync_slots: pd.DataFrame, block: str) -> list[list[dict[str, object]]]:
    block_df = sync_slots[sync_slots["Block"] == block].copy().sort_values(["Start", "End", "Course Codes", "Room"])
    if block_df.empty:
        return [[]]
    groups: list[list[dict[str, object]]] = []
    for _, group in block_df.groupby(["Start", "End"], sort=True):
        groups.append(group.to_dict("records"))
    return groups or [[]]


def _write_schedule_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    slots: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    report_title: str,
) -> None:
    rooms_present = sorted(slots["Room"].dropna().unique(), key=_room_sort_key) if not slots.empty else []
    rooms = [room for room in ROOM_ORDER if room in rooms_present]
    floor_a = [r for r in ["N108", "N110", "N135", "N150", "S115", "S120"] if r in rooms]
    floor_b = [r for r in ["N215", "N235", "N310"] if r in rooms]
    floor_c = [r for r in ["N335", "S340"] if r in rooms]
    floors = [("LOWER FLOOR", floor_a), ("UPPER FLOOR", floor_b), ("THIRD FLOOR", floor_c)]
    book = writer.book
    sheet = book.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = sheet
    sheet.hide_gridlines(2)
    dark = "#1F4E78"
    thin_gray = "#B7C9D6"
    border_gray = "#808080"
    title_fmt = book.add_format({"bold": True, "font_size": 15, "font_color": "white", "align": "center", "valign": "vcenter", "bg_color": dark})
    subtitle_fmt = book.add_format({"italic": True, "font_size": 8, "font_color": "#666666", "align": "center", "valign": "vcenter", "bg_color": "#F7FAFC"})
    floor_fmt = book.add_format({"bold": True, "font_size": 10, "align": "center", "valign": "vcenter", "top": 2, "top_color": dark, "bottom": 1, "bottom_color": thin_gray, "left": 1, "left_color": thin_gray, "right": 1, "right_color": thin_gray})
    room_header_blank_fmt = book.add_format({"bg_color": TIMEBLOCK_FILL, "top": 2, "top_color": border_gray, "bottom": 2, "bottom_color": dark, "left": 1, "left_color": thin_gray, "right": 1, "right_color": thin_gray})
    room_header_fmt = book.add_format({"bold": True, "font_size": 10, "font_color": "black", "align": "center", "valign": "vcenter", "bg_color": ROOM_HEADER_FILL, "top": 2, "top_color": dark, "bottom": 2, "bottom_color": dark, "left": 1, "left_color": thin_gray, "right": 1, "right_color": thin_gray})
    label_fmt = book.add_format({"bold": True, "font_size": 9, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TIMEBLOCK_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})
    special_label_fmt = book.add_format({"bold": True, "font_size": 9, "font_color": "black", "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": SPECIAL_ROW_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})
    special_note_top_fmt = book.add_format({"italic": True, "font_size": 9, "font_color": "black", "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": SPECIAL_ROW_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})
    term_header_fmt = book.add_format({"bold": True, "font_size": 10, "font_color": "white", "align": "center", "valign": "vcenter", "bg_color": TERM_HEADER_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})
    separator_blank_fmts: list[object] = []
    room_cell_fmts_by_col: list[dict[str, dict[str, object]]] = []
    for room_idx in range(len(rooms)):
        left_border = 2 if room_idx == 0 else 0
        separator_blank_fmts.append(
            book.add_format({
                "bg_color": "white",
                "align": "center",
                "valign": "vcenter",
                "left": left_border,
                "left_color": border_gray,
                "right": 2,
                "right_color": border_gray,
            })
        )
        subject_map: dict[str, dict[str, object]] = {}
        for subj, fill in SUBJECT_FILLS.items():
            subject_map[subj] = {
                "course": book.add_format({"font_size": 8, "bold": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": fill, "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "top": 2, "top_color": border_gray}),
                "title_none": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
                "title_sync": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC, "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
                "title_sync_async": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC_ASYNC, "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
                "details": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
                "annotation": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
                "annotation_prereq": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "#ffff00", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
            }
        room_cell_fmts_by_col.append(subject_map)
    default_fmts_by_col: list[dict[str, object]] = []
    for room_idx in range(len(rooms)):
        left_border = 2 if room_idx == 0 else 0
        default_fmts_by_col.append({
            "course": book.add_format({"font_size": 8, "bold": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "top": 2, "top_color": border_gray}),
            "title_none": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
            "title_sync": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC, "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
            "title_sync_async": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC_ASYNC, "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
            "details": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray}),
            "annotation": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
            "annotation_prereq": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "#ffff00", "left": left_border, "left_color": border_gray, "right": 2, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
        })

    timeblock_header_fmt = book.add_format({"bold": True, "font_size": 10, "font_color": "black", "align": "center", "valign": "vcenter", "bg_color": TIMEBLOCK_FILL, "top": 2, "top_color": dark, "bottom": 2, "bottom_color": dark, "left": 1, "left_color": thin_gray, "right": 1, "right_color": thin_gray})
    sheet.set_column(0, 0, 24)
    for i in range(1, len(rooms) + 1):
        sheet.set_column(i, i, 26)
    last_col = max(len(rooms), 1)
    sheet.merge_range(0, 0, 0, last_col, report_title, title_fmt)
    sheet.set_row(0, 24)
    sheet.merge_range(1, 0, 1, last_col, _printed_on_text(), subtitle_fmt)
    sheet.set_row(1, 18)
    sheet.write(2, 0, "Time Block", timeblock_header_fmt)
    col = 1
    for floor_name, floor_rooms in floors:
        if not floor_rooms:
            continue
        start_col = col
        end_col = col + len(floor_rooms) - 1
        if start_col == end_col:
            sheet.write(2, start_col, floor_name, floor_fmt)
        else:
            sheet.merge_range(2, start_col, 2, end_col, floor_name, floor_fmt)
        col = end_col + 1
    sheet.set_row(2, 18)
    sheet.write(3, 0, "", room_header_blank_fmt)
    for index, room in enumerate(rooms, start=1):
        sheet.write(3, index, _display_room(room), room_header_fmt)
    sheet.set_row(3, 18)
    current_row = 4
    for block in block_order:
        if block.endswith(" Header"):
            summary_text = term_summary_map.get(block, block.removesuffix(" Header"))
            if last_col >= 1:
                sheet.merge_range(current_row, 0, current_row, last_col, summary_text, term_header_fmt)
            else:
                sheet.write(current_row, 0, summary_text, term_header_fmt)
            sheet.set_row(current_row, 22)
            current_row += 1
            continue
        if block in special_rows:
            sheet.write(current_row, 0, block, special_label_fmt)
            if last_col > 1:
                sheet.merge_range(current_row, 1, current_row, last_col, special_rows[block], special_note_top_fmt)
            else:
                sheet.write(current_row, 1, special_rows[block], special_note_top_fmt)
            sheet.set_row(current_row, 22)
            current_row += 1
            continue
        block_df = slots[slots["Block"] == block].copy().sort_values(["Start", "Room", "Course Codes"]) if not slots.empty else pd.DataFrame()
        room_lists = {room: block_df[block_df["Room"] == room].to_dict("records") for room in rooms}
        max_len = max([len(v) for v in room_lists.values()] + [1])
        start_row = current_row
        end_row = current_row + (max_len * 4) - 1
        if end_row > start_row:
            sheet.merge_range(start_row, 0, end_row, 0, block, label_fmt)
        else:
            sheet.write(start_row, 0, block, label_fmt)
        top_band_fmt = book.add_format({"top": 2, "top_color": border_gray})
        for col_index in range(1, len(rooms) + 1):
            sheet.write_blank(start_row, col_index, "", top_band_fmt)
        for slot_index in range(max_len):
            course_row = start_row + (slot_index * 4)
            title_row = course_row + 1
            details_row = course_row + 2
            annotation_row = course_row + 3
            sheet.set_row(course_row, 22)
            sheet.set_row(title_row, 26)
            sheet.set_row(details_row, 40)
            sheet.set_row(annotation_row, 20)
            for col_index, room in enumerate(rooms, start=1):
                recs = room_lists[room]
                if slot_index < len(recs):
                    rec = recs[slot_index]
                    subj = _course_subject_from_key(_first_course_key(rec.get("Course Codes")))
                    fmts = room_cell_fmts_by_col[col_index - 1].get(subj, default_fmts_by_col[col_index - 1])
                    _write_stacked_cells(sheet, course_row, title_row, details_row, annotation_row, col_index, rec, fmts, title_method_map, course_annotation_map)
                else:
                    blank_fmt = separator_blank_fmts[col_index - 1]
                    sheet.write_blank(course_row, col_index, "", blank_fmt)
                    sheet.write_blank(title_row, col_index, "", blank_fmt)
                    sheet.write_blank(details_row, col_index, "", blank_fmt)
                    sheet.write_blank(annotation_row, col_index, "", blank_fmt)
        current_row = end_row + 1
    sheet.freeze_panes(4, 1)
    sheet.repeat_rows(0, 3)
    sheet.set_landscape()
    sheet.set_paper(1)
    sheet.fit_to_pages(1, 1)
    sheet.set_margins(left=0.2, right=0.2, top=0.3, bottom=0.3)
    sheet.print_area(0, 0, current_row - 1, last_col)


def _write_sync_schedule_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    sync_slots: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    report_title: str,
) -> None:
    display_blocks = _sync_display_blocks(block_order, special_rows)
    slot_counts = [
        len(sync_slots[sync_slots["Block"] == block].index)
        for block in display_blocks
        if not block.endswith(" Header")
    ]
    max_slots = max(slot_counts + [1])
    book = writer.book
    sheet = book.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = sheet
    sheet.hide_gridlines(2)

    dark = "#1F4E78"
    thin_gray = "#B7C9D6"
    border_gray = "#808080"
    title_fmt = book.add_format({"bold": True, "font_size": 15, "font_color": "white", "align": "center", "valign": "vcenter", "bg_color": dark})
    subtitle_fmt = book.add_format({"italic": True, "font_size": 8, "font_color": "#666666", "align": "center", "valign": "vcenter", "bg_color": "#F7FAFC"})
    label_fmt = book.add_format({"bold": True, "font_size": 11, "align": "center", "valign": "vcenter", "text_wrap": True, "rotation": 90, "bg_color": TIMEBLOCK_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})
    term_header_fmt = book.add_format({"bold": True, "font_size": 10, "font_color": "white", "align": "center", "valign": "vcenter", "bg_color": TERM_HEADER_FILL, "border": 1, "border_color": thin_gray, "top": 2, "top_color": border_gray})

    slot_fmts_by_col: list[dict[str, dict[str, object]]] = []
    blank_fmts_by_col: list[dict[str, object]] = []
    for slot_idx in range(max_slots):
        left_border = 2 if slot_idx == 0 else 1
        blank_fmts_by_col.append({
            "course": book.add_format({"bg_color": "white"}),
            "title": book.add_format({"bg_color": "white"}),
            "details": book.add_format({"bg_color": "white"}),
        })
        subject_map: dict[str, dict[str, object]] = {}
        for subj, fill in SUBJECT_FILLS.items():
            subject_map[subj] = {
                "course": book.add_format({"font_size": 8, "bold": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": fill, "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "top": 2, "top_color": border_gray}),
                "title_none": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
                "title_sync": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC, "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
                "title_sync_async": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC_ASYNC, "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
                "details": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
                "annotation": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
                "annotation_prereq": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "#ffff00", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
            }
        slot_fmts_by_col.append(subject_map)

    default_fmts_by_col: list[dict[str, object]] = []
    for slot_idx in range(max_slots):
        left_border = 2 if slot_idx == 0 else 1
        default_fmts_by_col.append({
            "course": book.add_format({"font_size": 8, "bold": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "top": 2, "top_color": border_gray}),
            "title_none": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
            "title_sync": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC, "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
            "title_sync_async": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": TITLE_FILL_SYNC_ASYNC, "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
            "details": book.add_format({"font_size": 8, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray}),
            "annotation": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "white", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
            "annotation_prereq": book.add_format({"font_size": 8, "italic": True, "align": "center", "valign": "vcenter", "text_wrap": True, "bg_color": "#ffff00", "left": left_border, "left_color": border_gray, "right": 1, "right_color": border_gray, "bottom": 2, "bottom_color": border_gray}),
        })

    sheet.set_column(0, 0, 16)
    for i in range(1, max_slots + 1):
        sheet.set_column(i, i, 28)
    last_col = max_slots
    sheet.merge_range(0, 0, 0, last_col, report_title, title_fmt)
    sheet.set_row(0, 24)
    sheet.merge_range(1, 0, 1, last_col, _printed_on_text(), subtitle_fmt)
    sheet.set_row(1, 18)

    current_row = 2
    for block in display_blocks:
        if block.endswith(" Header"):
            summary_text = term_summary_map.get(block, block.removesuffix(" Header"))
            sheet.merge_range(current_row, 0, current_row, last_col, summary_text, term_header_fmt)
            sheet.set_row(current_row, 22)
            current_row += 1
            continue

        time_groups = _sync_block_time_groups(sync_slots, block)
        block_start_row = current_row
        block_end_row = current_row + (len(time_groups) * 4) - 1
        sheet.merge_range(block_start_row, 0, block_end_row, 0, block, label_fmt)

        for recs in time_groups:
            course_row = current_row
            title_row = current_row + 1
            details_row = current_row + 2
            annotation_row = current_row + 3
            sheet.set_row(course_row, 22)
            sheet.set_row(title_row, 26)
            sheet.set_row(details_row, 78)
            sheet.set_row(annotation_row, 20)

            for slot_index in range(max_slots):
                col_index = slot_index + 1
                if slot_index < len(recs):
                    rec = recs[slot_index]
                    subj = _course_subject_from_key(_first_course_key(rec.get("Course Codes")))
                    fmts = slot_fmts_by_col[slot_index].get(subj, default_fmts_by_col[slot_index])
                    _write_stacked_cells(sheet, course_row, title_row, details_row, annotation_row, col_index, rec, fmts, title_method_map, course_annotation_map)
                else:
                    blank_fmts = blank_fmts_by_col[slot_index]
                    sheet.write_blank(course_row, col_index, "", blank_fmts["course"])
                    sheet.write_blank(title_row, col_index, "", blank_fmts["title"])
                    sheet.write_blank(details_row, col_index, "", blank_fmts["details"])
                    sheet.write_blank(annotation_row, col_index, "", blank_fmts["details"])
            current_row += 4

    sheet.freeze_panes(2, 1)
    sheet.repeat_rows(0, 1)
    sheet.set_landscape()
    sheet.set_paper(1)
    sheet.fit_to_pages(1, 0)
    sheet.set_margins(left=0.2, right=0.2, top=0.3, bottom=0.3)
    sheet.print_area(0, 0, current_row - 1, last_col)


def _build_schedule_excel(
    slots: pd.DataFrame,
    sync_slots: pd.DataFrame,
    meetings: pd.DataFrame,
    excluded_df: pd.DataFrame,
    async_report_df: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    sync_block_order: list[str],
    sync_special_rows: dict[str, str],
    sync_term_summary_map: dict[str, str],
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        _write_schedule_sheet(
            writer,
            "Weekly Schedule",
            slots,
            title_method_map,
            course_annotation_map,
            block_order,
            special_rows,
            term_summary_map,
            "Printable Weekly Schedule - In Person Classes",
        )
        _write_sync_schedule_sheet(
            writer,
            "SYNC Schedule",
            sync_slots,
            title_method_map,
            course_annotation_map,
            sync_block_order,
            sync_special_rows,
            sync_term_summary_map,
            "Printable Weekly Schedule - Sync Classes",
        )
        _write_async_schedule_sheet(
            writer,
            "ASYNC Classes",
            async_report_df,
            title_method_map,
            "Printable Weekly Schedule - Async Classes",
        )
        meetings.to_excel(writer, sheet_name="Normalized Meetings", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded", index=False)
    output.seek(0)
    return output.read()


def _title_format_key(course_codes: object, title: str, title_method_map: dict[str, set[str]]) -> str:
    fill = _title_fill_for_methods(_method_set_for_slot(course_codes, title, title_method_map))
    if fill == TITLE_FILL_SYNC_ASYNC:
        return "title_sync_async"
    if fill == TITLE_FILL_SYNC:
        return "title_sync"
    return "title_none"


def _write_stacked_cells(
    ws,
    course_row: int,
    title_row: int,
    details_row: int,
    annotation_row: int,
    col: int,
    rec: dict[str, object],
    fmts: dict[str, object],
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
) -> None:
    title_text = str(rec.get("Title") or "").strip()
    ws.write(course_row, col, str(rec.get("Course Codes") or "").strip(), fmts["course"])
    ws.write(title_row, col, title_text, fmts[_title_format_key(rec.get("Course Codes"), title_text, title_method_map)])
    credit_lines = rec.get("Credit CRN Lines") or []
    if not isinstance(credit_lines, list):
        credit_lines = [str(credit_lines)] if credit_lines else []
    details_lines = [line for line in [
        str(rec.get("Instructor") or "").strip(),
        *credit_lines,
        str(rec.get("DayTime") or "").strip(),
    ] if line]
    ws.write(details_row, col, "\n".join(details_lines), fmts["details"])
    annotation_text = _course_annotation_for_slot(rec.get("Course Codes"), course_annotation_map)
    annotation_fmt = fmts["annotation_prereq"] if _is_prerequisite_annotation(annotation_text) else fmts["annotation"]
    ws.write(annotation_row, col, annotation_text, annotation_fmt)


def _write_async_schedule_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    async_report_df: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    report_title: str,
) -> None:
    book = writer.book
    sheet = book.add_worksheet(sheet_name)
    writer.sheets[sheet_name] = sheet
    sheet.hide_gridlines(2)

    dark = "#1F4E78"
    border_gray = "#808080"
    title_fmt = book.add_format({"bold": True, "font_size": 15, "font_color": "white", "align": "center", "valign": "vcenter", "bg_color": dark})
    subtitle_fmt = book.add_format({"italic": True, "font_size": 8, "font_color": "#666666", "align": "center", "valign": "vcenter", "bg_color": "#F7FAFC"})
    header_fmt = book.add_format({"bold": True, "font_size": 10, "align": "center", "valign": "vcenter", "bg_color": "#EDEDED", "border": 1, "border_color": border_gray})
    cell_fmt = book.add_format({"font_size": 9, "align": "center", "valign": "vcenter", "border": 1, "border_color": border_gray})
    text_left_fmt = book.add_format({"font_size": 9, "align": "left", "valign": "vcenter", "border": 1, "border_color": border_gray})
    instructor_left_fmt = book.add_format({"font_size": 9, "align": "left", "valign": "vcenter", "border": 1, "border_color": border_gray})
    prereq_fmt = book.add_format({"font_size": 9, "align": "left", "valign": "vcenter", "bg_color": "#ffff00", "border": 1, "border_color": border_gray})

    subject_cell_fmts = {
        subj: book.add_format({"font_size": 9, "align": "center", "valign": "vcenter", "bg_color": fill, "border": 1, "border_color": border_gray})
        for subj, fill in SUBJECT_FILLS.items()
    }
    title_sync_fmt = book.add_format({"font_size": 9, "align": "left", "valign": "vcenter", "bg_color": TITLE_FILL_SYNC, "border": 1, "border_color": border_gray})
    title_sync_async_fmt = book.add_format({"font_size": 9, "align": "left", "valign": "vcenter", "bg_color": TITLE_FILL_SYNC_ASYNC, "border": 1, "border_color": border_gray})

    columns = ["Course / Sect.", "CRN #", "Course Title", "Attribute/Notes", "Cr.", "Instructor", "Cap"]
    widths = [18, 10, 30, 28, 7, 18, 8]
    for idx, width in enumerate(widths):
        sheet.set_column(idx, idx, width)
    last_col = len(columns) - 1
    sheet.merge_range(0, 0, 0, last_col, report_title, title_fmt)
    sheet.merge_range(1, 0, 1, last_col, _printed_on_text(), subtitle_fmt)
    sheet.set_row(0, 24)
    sheet.set_row(1, 18)
    for col_idx, name in enumerate(columns):
        sheet.write(2, col_idx, name, header_fmt)
    sheet.set_row(2, 20)

    for row_idx, (_, row) in enumerate(async_report_df.iterrows(), start=3):
        subject_fmt = subject_cell_fmts.get(str(row.get("Subject Code") or "").strip(), cell_fmt)
        title_fill = _title_fill_for_methods(_method_set_for_slot(row.get("Course Key"), row.get("Course Title"), title_method_map))
        title_fmt_for_row = title_sync_async_fmt if title_fill == TITLE_FILL_SYNC_ASYNC else title_sync_fmt if title_fill == TITLE_FILL_SYNC else text_left_fmt
        attr_text = row["Attribute/Notes"]
        attr_fmt = prereq_fmt if _is_prerequisite_annotation(attr_text) else text_left_fmt
        sheet.write(row_idx, 0, row["Course / Sect."], subject_fmt)
        sheet.write(row_idx, 1, row["CRN #"], cell_fmt)
        sheet.write(row_idx, 2, row["Course Title"], title_fmt_for_row)
        sheet.write(row_idx, 3, attr_text, attr_fmt)
        sheet.write(row_idx, 4, row["Cr."], cell_fmt)
        sheet.write(row_idx, 5, row["Instructor"], instructor_left_fmt)
        sheet.write(row_idx, 6, row["Cap"], cell_fmt)

    sheet.freeze_panes(3, 0)
    sheet.repeat_rows(0, 2)
    sheet.set_landscape()
    sheet.set_paper(1)
    sheet.fit_to_pages(1, 0)
    sheet.set_margins(left=0.2, right=0.2, top=0.3, bottom=0.3)
    sheet.print_area(0, 0, max(len(async_report_df) + 2, 2), last_col)


def _format_credit_value(credits: str) -> str:
    credits = str(credits or "").strip()
    if not credits:
        return ""
    if credits.endswith(".0"):
        credits = credits[:-2]
    return f"{credits} Cr."


def _format_credit_crn(credits: str, crns: str, enrollment_cap: str = "") -> str:
    credit_text = _format_credit_value(credits)
    crns = str(crns or "").strip()
    enrollment_text = _normalize_enrollment(enrollment_cap)
    if credit_text and crns and enrollment_text:
        return f"{credit_text} (CRN {crns}) {enrollment_text} St."
    if credit_text and crns:
        return f"{credit_text} (CRN {crns})"
    if credit_text and enrollment_text:
        return f"{credit_text} {enrollment_text} St."
    if credit_text:
        return credit_text
    if crns and enrollment_text:
        return f"(CRN {crns}) {enrollment_text} St."
    if crns:
        return f"(CRN {crns})"
    if enrollment_text:
        return f"{enrollment_text} St."
    return ""


def _build_credit_crn_lines(group: pd.DataFrame) -> list[str]:
    if group.empty:
        return []
    unique_codes = _uniq(group["Course Code"])
    if len(unique_codes) <= 1:
        row = group.iloc[0]
        line = _format_credit_crn(
            str(row.get("Credits") or ""),
            str(row.get("CRN") or ""),
            str(row.get("Enrollment Cap") or ""),
        )
        return [line] if line else []

    family_counts: dict[str, int] = {}
    for course_code in unique_codes:
        family = _course_family_from_key(course_code)
        family_counts[family] = family_counts.get(family, 0) + 1

    lines: list[str] = []
    seen: set[str] = set()
    for course_code in unique_codes:
        course_rows = group[group["Course Code"].astype(str) == str(course_code)].reset_index(drop=True)
        if course_rows.empty:
            continue
        row = course_rows.iloc[0]
        family = _course_family_from_key(course_code)
        label = course_code if family_counts.get(family, 0) > 1 else family
        line_body = _format_credit_crn(
            str(row.get("Credits") or ""),
            str(row.get("CRN") or ""),
            str(row.get("Enrollment Cap") or ""),
        )
        line = f"{label} - {line_body}".strip(" -")
        if line and line not in seen:
            seen.add(line)
            lines.append(line)
    return lines


def _build_schedule_pdf_elements(
    slots: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    *,
    title_text: str,
    subtitle_text: str,
    add_bottom_rule: bool = False,
) -> list[object]:
    page_size = landscape(A2)
    rooms_present = sorted(slots["Room"].dropna().unique(), key=_room_sort_key) if not slots.empty else []
    rooms = [room for room in ROOM_ORDER if room in rooms_present]
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1F4E78"), alignment=1, spaceAfter=6)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#666666"), alignment=1, spaceAfter=8)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.2, leading=7, spaceAfter=0, alignment=1)
    annotation_style = ParagraphStyle("annotation_cell", parent=styles["Normal"], fontSize=6.2, leading=7, spaceAfter=0, alignment=1, fontName="Helvetica-Oblique")
    label_style = ParagraphStyle("label", parent=styles["Normal"], fontSize=7, leading=8, alignment=1)
    special_style = ParagraphStyle("special", parent=styles["Normal"], fontSize=7, leading=8, alignment=1, textColor=colors.black)
    term_header_style = ParagraphStyle("term_header", parent=styles["Normal"], fontSize=7, leading=8, alignment=1, textColor=colors.white)

    data: list[list[object]] = []
    header_row: list[object] = [Paragraph("Time Block", label_style)]
    for room in rooms:
        header_row.append(Paragraph(_display_room(room), label_style))
    data.append(header_row)
    row_styles: list[tuple[int, str]] = [(0, "header")]
    block_start_rows: list[int] = []
    class_boxes: list[tuple[int, int]] = []
    for block in block_order:
        if block.endswith(" Header"):
            data.append([Paragraph(term_summary_map.get(block, block.removesuffix(" Header")), term_header_style), *["" for _ in range(max(len(rooms), 0))]])
            row_styles.append((len(data) - 1, "term_header"))
            continue
        if block in special_rows:
            data.append([Paragraph(block, label_style), Paragraph(special_rows[block], special_style), *["" for _ in range(max(len(rooms) - 1, 0))]])
            row_styles.append((len(data) - 1, "special"))
            continue
        block_df = slots[slots["Block"] == block].copy().sort_values(["Start", "Room", "Course Codes"]) if not slots.empty else pd.DataFrame()
        room_lists = {room: block_df[block_df["Room"] == room].to_dict("records") for room in rooms}
        max_len = max([len(v) for v in room_lists.values()] + [1])
        for slot_index in range(max_len):
            if slot_index == 0:
                block_start_rows.append(len(data))
            label_cell = Paragraph(block if slot_index == 0 else "", label_style)
            course_cells: list[object] = [label_cell]
            title_cells: list[object] = [""]
            detail_cells: list[object] = [""]
            annotation_cells: list[object] = [""]
            for room in rooms:
                recs = room_lists[room]
                if slot_index < len(recs):
                    code_para, title_para, details_para, annotation_para = _schedule_paragraph_cells(recs[slot_index], cell_style, annotation_style, title_method_map, course_annotation_map)
                    course_cells.append(code_para)
                    title_cells.append(title_para)
                    detail_cells.append(details_para)
                    annotation_cells.append(annotation_para)
                else:
                    course_cells.append("")
                    title_cells.append("")
                    detail_cells.append("")
                    annotation_cells.append("")
            data.extend([course_cells, title_cells, detail_cells, annotation_cells])
            row_styles.extend([(len(data) - 4, "course"), (len(data) - 3, "title"), (len(data) - 2, "details"), (len(data) - 1, "annotation")])
            for room_col in range(1, len(rooms) + 1):
                if slot_index < len(room_lists[rooms[room_col - 1]]):
                    class_boxes.append((len(data) - 4, room_col))

    if not data:
        data = [["No schedule rows found"]]

    available_width = page_size[0] - 24
    first_col = 95
    remaining = max(len(rooms), 1)
    other_col = max((available_width - first_col) / remaining, 55)
    col_widths = [first_col] + [other_col for _ in rooms]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(TIMEBLOCK_FILL)),
        ("BACKGROUND", (1, 0), (-1, 0), colors.HexColor(ROOM_HEADER_FILL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.HexColor("#B7C9D6")),
    ]
    for row_index, row_kind in row_styles[1:]:
        if row_kind == "term_header":
            style_cmds.extend([
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(TERM_HEADER_FILL)),
                ("TEXTCOLOR", (0, row_index), (-1, row_index), colors.white),
                ("SPAN", (0, row_index), (-1, row_index)),
                ("LINEABOVE", (0, row_index), (-1, row_index), 1.5, colors.HexColor("#808080")),
            ])
        elif row_kind == "special":
            style_cmds.extend([
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(SPECIAL_ROW_FILL)),
                ("SPAN", (1, row_index), (-1, row_index)),
                ("LINEABOVE", (0, row_index), (-1, row_index), 1.5, colors.HexColor("#808080")),
            ])
        elif row_kind in {"course", "title", "details", "annotation"}:
            style_cmds.append(("BACKGROUND", (0, row_index), (0, row_index), colors.HexColor(TIMEBLOCK_FILL)))
            if row_kind == "annotation":
                style_cmds.append(("TOPPADDING", (1, row_index), (-1, row_index), 3))
                style_cmds.append(("BOTTOMPADDING", (1, row_index), (-1, row_index), 4))
    for room_col in range(1, len(rooms) + 1):
        style_cmds.append(("LINEAFTER", (room_col, 0), (room_col, -1), 1.5, colors.HexColor("#808080")))
    if rooms:
        style_cmds.append(("LINEBEFORE", (1, 0), (1, -1), 1.5, colors.HexColor("#808080")))
    style_cmds.append(("LINEABOVE", (0, 0), (-1, 0), 1.5, colors.HexColor("#808080")))
    for row_index in block_start_rows:
        style_cmds.append(("LINEABOVE", (0, row_index), (-1, row_index), 1.5, colors.HexColor("#808080")))
    for start_row, col in class_boxes:
        style_cmds.append(("BOX", (col, start_row), (col, start_row + 3), 1.5, colors.HexColor("#808080")))
    for ridx in range(1, len(data)):
        row_kind = row_styles[ridx][1] if ridx < len(row_styles) else ""
        for cidx in range(1, len(data[ridx])):
            rec = getattr(data[ridx][cidx], "_schedule_record", None)
            if not rec:
                continue
            if row_kind == "course":
                subj = _course_subject_from_key(_first_course_key(rec.get("Course Codes", "")))
                style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor(SUBJECT_FILLS.get(subj, "#FFFFFF"))))
            elif row_kind == "title":
                title_fill = _title_fill_for_methods(_method_set_for_slot(rec.get("Course Codes"), rec.get("Title"), title_method_map))
                if title_fill:
                    style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor(title_fill)))
            elif row_kind == "annotation":
                annotation_text = _course_annotation_for_slot(rec.get("Course Codes"), course_annotation_map)
                if _is_prerequisite_annotation(annotation_text):
                    style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor("#ffff00")))
    if add_bottom_rule and data:
        style_cmds.append(("LINEBELOW", (0, len(data) - 1), (-1, len(data) - 1), 1.5, colors.HexColor("#808080")))
    table.setStyle(TableStyle(style_cmds))
    return [
        Paragraph(title_text, title_style),
        Paragraph(subtitle_text, subtitle_style),
        table,
        Spacer(1, 8),
    ]


def _build_sync_schedule_pdf_elements(
    sync_slots: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    *,
    title_text: str,
    subtitle_text: str,
) -> list[object]:
    page_size = landscape(A2)
    display_blocks = _sync_display_blocks(block_order, special_rows)
    slot_counts = [
        len(sync_slots[sync_slots["Block"] == block].index)
        for block in display_blocks
        if not block.endswith(" Header")
    ]
    max_slots = max(slot_counts + [1])

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("sync_title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1F4E78"), alignment=1, spaceAfter=6)
    subtitle_style = ParagraphStyle("sync_subtitle", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#666666"), alignment=1, spaceAfter=8)
    cell_style = ParagraphStyle("sync_cell", parent=styles["Normal"], fontSize=8, leading=9, spaceAfter=0, alignment=1)
    annotation_style = ParagraphStyle("sync_annotation_cell", parent=styles["Normal"], fontSize=8, leading=9, spaceAfter=0, alignment=1, fontName="Helvetica-Oblique")
    label_style = ParagraphStyle("sync_label", parent=styles["Normal"], fontSize=10, leading=11, alignment=1)
    term_header_style = ParagraphStyle("sync_term_header", parent=styles["Normal"], fontSize=8, leading=9, alignment=1, textColor=colors.white)

    data: list[list[object]] = []
    row_kind_by_index: dict[int, str] = {}
    block_spans: list[tuple[int, int]] = []
    for block in display_blocks:
        if block.endswith(" Header"):
            data.append([Paragraph(term_summary_map.get(block, block.removesuffix(" Header")), term_header_style), *["" for _ in range(max_slots)]])
            row_kind_by_index[len(data) - 1] = "term_header"
            continue

        time_groups = _sync_block_time_groups(sync_slots, block)
        block_start_row = len(data)
        for group_index, recs in enumerate(time_groups):
            label_cell = Paragraph(block.replace(" ", "<br/>") if group_index == 0 else "", label_style)
            course_cells: list[object] = [label_cell]
            title_cells: list[object] = [""]
            detail_cells: list[object] = [""]
            annotation_cells: list[object] = [""]
            for slot_index in range(max_slots):
                if slot_index < len(recs):
                    code_para, title_para, details_para, annotation_para = _schedule_paragraph_cells(recs[slot_index], cell_style, annotation_style, title_method_map, course_annotation_map)
                    course_cells.append(code_para)
                    title_cells.append(title_para)
                    detail_cells.append(details_para)
                    annotation_cells.append(annotation_para)
                else:
                    course_cells.append("")
                    title_cells.append("")
                    detail_cells.append("")
                    annotation_cells.append("")
            data.extend([course_cells, title_cells, detail_cells, annotation_cells])
            row_kind_by_index[len(data) - 4] = "course"
            row_kind_by_index[len(data) - 3] = "title"
            row_kind_by_index[len(data) - 2] = "details"
            row_kind_by_index[len(data) - 1] = "annotation"
        block_end_row = len(data) - 1
        block_spans.append((block_start_row, block_end_row))

    if not data:
        data = [["No SYNC schedule rows found"]]

    available_width = page_size[0] - 24
    first_col = 72
    other_col = max((available_width - first_col) / max_slots, 110)
    col_widths = [first_col] + [other_col for _ in range(max_slots)]
    table = Table(data, colWidths=col_widths)
    style_cmds = [("VALIGN", (0, 0), (-1, -1), "TOP")]

    for row_index, row_kind in row_kind_by_index.items():
        if row_kind == "term_header":
            style_cmds.extend([
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(TERM_HEADER_FILL)),
                ("TEXTCOLOR", (0, row_index), (-1, row_index), colors.white),
                ("SPAN", (0, row_index), (-1, row_index)),
                ("LINEABOVE", (0, row_index), (-1, row_index), 1.5, colors.HexColor("#808080")),
            ])
            continue
        if row_kind in {"course", "title", "details", "annotation"}:
            style_cmds.append(("BACKGROUND", (0, row_index), (0, row_index), colors.HexColor(TIMEBLOCK_FILL)))
            style_cmds.append(("ALIGN", (0, row_index), (0, row_index), "CENTER"))
            style_cmds.append(("VALIGN", (0, row_index), (0, row_index), "MIDDLE"))
            if row_kind == "annotation":
                style_cmds.append(("TOPPADDING", (1, row_index), (-1, row_index), 3))
                style_cmds.append(("BOTTOMPADDING", (1, row_index), (-1, row_index), 4))
    for block_start, block_end in block_spans:
        style_cmds.append(("SPAN", (0, block_start), (0, block_end)))
        style_cmds.append(("BOX", (0, block_start), (0, block_end), 1.5, colors.HexColor("#808080")))
    for ridx in range(len(data)):
        row_kind = row_kind_by_index.get(ridx, "")
        if row_kind == "term_header":
            continue
        if row_kind in {"course", "title", "details", "annotation"}:
            for cidx in range(1, len(data[ridx])):
                rec = getattr(data[ridx][cidx], "_schedule_record", None)
                if not rec:
                    continue
                top_width = 1.5 if row_kind == "course" else 0 if row_kind == "annotation" else 0.35
                bottom_width = 1.5 if row_kind == "annotation" else 0 if row_kind == "details" else 0.35
                style_cmds.extend([
                    ("LINEBEFORE", (cidx, ridx), (cidx, ridx), 1.0, colors.HexColor("#808080")),
                    ("LINEAFTER", (cidx, ridx), (cidx, ridx), 1.0, colors.HexColor("#808080")),
                ])
                if top_width > 0:
                    style_cmds.append(("LINEABOVE", (cidx, ridx), (cidx, ridx), top_width, colors.HexColor("#808080")))
                if bottom_width > 0:
                    style_cmds.append(("LINEBELOW", (cidx, ridx), (cidx, ridx), bottom_width, colors.HexColor("#808080")))
                if row_kind == "course":
                    subj = _course_subject_from_key(_first_course_key(rec.get("Course Codes", "")))
                    style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor(SUBJECT_FILLS.get(subj, "#FFFFFF"))))
                elif row_kind == "title":
                    title_fill = _title_fill_for_methods(_method_set_for_slot(rec.get("Course Codes"), rec.get("Title"), title_method_map))
                    if title_fill:
                        style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor(title_fill)))
                elif row_kind == "annotation":
                    annotation_text = _course_annotation_for_slot(rec.get("Course Codes"), course_annotation_map)
                    if _is_prerequisite_annotation(annotation_text):
                        style_cmds.append(("BACKGROUND", (cidx, ridx), (cidx, ridx), colors.HexColor("#ffff00")))

    table.setStyle(TableStyle(style_cmds))
    return [
        Paragraph(title_text, title_style),
        Paragraph(subtitle_text, subtitle_style),
        table,
        Spacer(1, 8),
    ]


def _build_async_schedule_pdf_elements(
    async_report_df: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    subtitle_text: str,
) -> list[object]:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("async_title", parent=styles["Title"], fontSize=16, textColor=colors.HexColor("#1F4E78"), alignment=1, spaceAfter=6)
    subtitle_style = ParagraphStyle("async_subtitle", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#666666"), alignment=1, spaceAfter=8)
    header_style = ParagraphStyle("async_header", parent=styles["Normal"], fontSize=8, leading=9, alignment=1)
    cell_style = ParagraphStyle("async_cell", parent=styles["Normal"], fontSize=8, leading=9, alignment=1)
    cell_left_style = ParagraphStyle("async_cell_left", parent=styles["Normal"], fontSize=8, leading=9, alignment=0)

    data: list[list[object]] = [[
        Paragraph("Course / Sect.", header_style),
        Paragraph("CRN #", header_style),
        Paragraph("Course Title", header_style),
        Paragraph("Attribute/Notes", header_style),
        Paragraph("Cr.", header_style),
        Paragraph("Instructor", header_style),
        Paragraph("Cap", header_style),
    ]]
    for _, row in async_report_df.iterrows():
        row_cells = [
            Paragraph(str(row["Course / Sect."] or ""), cell_style),
            Paragraph(str(row["CRN #"] or ""), cell_style),
            Paragraph(str(row["Course Title"] or "").replace("&", "&amp;"), cell_left_style),
            Paragraph(str(row["Attribute/Notes"] or "").replace("&", "&amp;"), cell_left_style),
            Paragraph(str(row["Cr."] or ""), cell_style),
            Paragraph(str(row["Instructor"] or "").replace("&", "&amp;"), cell_left_style),
            Paragraph(str(row["Cap"] or ""), cell_style),
        ]
        for cell in row_cells:
            setattr(cell, "_async_row", row.to_dict())
        data.append(row_cells)
    if len(data) == 1:
        data.append([Paragraph("", cell_style) for _ in range(7)])

    table = Table(data, colWidths=[90, 55, 170, 150, 40, 90, 45], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EDEDED")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#808080")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for ridx in range(1, len(data)):
        row_meta = getattr(data[ridx][0], "_async_row", None)
        if not row_meta:
            continue
        subj = str(row_meta.get("Subject Code") or "").strip()
        if subj in SUBJECT_FILLS:
            style_cmds.append(("BACKGROUND", (0, ridx), (0, ridx), colors.HexColor(SUBJECT_FILLS[subj])))
        title_fill = _title_fill_for_methods(_method_set_for_slot(row_meta.get("Course Key"), row_meta.get("Course Title"), title_method_map))
        if title_fill:
            style_cmds.append(("BACKGROUND", (2, ridx), (2, ridx), colors.HexColor(title_fill)))
        if _is_prerequisite_annotation(row_meta.get("Attribute/Notes")):
            style_cmds.append(("BACKGROUND", (3, ridx), (3, ridx), colors.HexColor("#ffff00")))
        style_cmds.append(("ALIGN", (5, ridx), (5, ridx), "LEFT"))
    table.setStyle(TableStyle(style_cmds))
    return [
        Paragraph("Printable Weekly Schedule - Async Classes", title_style),
        Paragraph(subtitle_text, subtitle_style),
        table,
        Spacer(1, 8),
    ]


def _build_schedule_pdf(
    slots: pd.DataFrame,
    sync_slots: pd.DataFrame,
    async_report_df: pd.DataFrame,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
    block_order: list[str],
    special_rows: dict[str, str],
    term_summary_map: dict[str, str],
    sync_block_order: list[str],
    sync_special_rows: dict[str, str],
    sync_term_summary_map: dict[str, str],
) -> bytes:
    pdf_buffer = BytesIO()
    page_size = landscape(A2)
    doc = SimpleDocTemplate(pdf_buffer, pagesize=page_size, leftMargin=12, rightMargin=12, topMargin=12, bottomMargin=12)
    elements = _build_schedule_pdf_elements(
        slots,
        title_method_map,
        course_annotation_map,
        block_order,
        special_rows,
        term_summary_map,
        title_text="Printable Weekly Schedule - In Person Classes",
        subtitle_text=_printed_on_text(),
        add_bottom_rule=True,
    )
    elements.append(PageBreak())
    elements.extend(
        _build_sync_schedule_pdf_elements(
            sync_slots,
            title_method_map,
            course_annotation_map,
            sync_block_order,
            sync_special_rows,
            sync_term_summary_map,
            title_text="Printable Weekly Schedule - Sync Classes",
            subtitle_text=_printed_on_text(),
        )
    )
    elements.append(PageBreak())
    elements.extend(_build_async_schedule_pdf_elements(async_report_df, title_method_map, _printed_on_text()))
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer.read()


def _schedule_paragraph_cells(
    rec: dict[str, object],
    style: ParagraphStyle,
    annotation_style: ParagraphStyle,
    title_method_map: dict[str, set[str]],
    course_annotation_map: dict[str, str],
) -> list[Paragraph]:
    title = str(rec.get("Title") or "").replace("&", "&amp;")
    codes = str(rec.get("Course Codes") or "").replace("&", "&amp;")
    instructor = str(rec.get("Instructor") or "").replace("&", "&amp;")
    credit_lines = rec.get("Credit CRN Lines") or []
    if not isinstance(credit_lines, list):
        credit_lines = [str(credit_lines)] if credit_lines else []
    safe_credit_lines = [str(line).replace("&", "&amp;") for line in credit_lines if str(line).strip()]
    day_time = str(rec.get("DayTime") or "").replace("&", "&amp;")
    annotation = _course_annotation_for_slot(rec.get("Course Codes"), course_annotation_map).replace("&", "&amp;")
    code_para = Paragraph(f"<b>{codes}</b>", style)
    title_para = Paragraph(title or "", style)
    details_text = "<br/>".join([piece for piece in [instructor, *safe_credit_lines, day_time] if piece])
    details_para = Paragraph(details_text, style)
    annotation_para = Paragraph(annotation or "", annotation_style)
    setattr(code_para, "_schedule_record", rec)
    setattr(title_para, "_schedule_record", rec)
    setattr(details_para, "_schedule_record", rec)
    setattr(annotation_para, "_schedule_record", rec)
    return [code_para, title_para, details_para, annotation_para]


def _safe_stem(source_name: str) -> str:
    stem = re.sub(r"\.xlsx$", "", source_name, flags=re.I)
    return re.sub(r"[^\w\-.]+", "_", stem).strip("_") or "donsheet"
